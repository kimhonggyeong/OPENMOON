from __future__ import annotations

import email
import imaplib
import os
import re
import shutil
import time

from datetime import datetime
from email.header import decode_header
from email.message import Message
from email.utils import parseaddr
from pathlib import Path
from typing import Literal

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from pydantic import BaseModel, Field


# =========================================================
# 환경 변수
# =========================================================

load_dotenv()

LOGIN_ID = os.getenv(
    "DAUM_LOGIN_ID",
    ""
).strip()

APP_PASSWORD = os.getenv(
    "DAUM_APP_PASSWORD",
    ""
).strip()

OPENAI_API_KEY = os.getenv(
    "OPENAI_API_KEY",
    ""
).strip()


# =========================================================
# 기본 설정
# =========================================================

IMAP_SERVER = "imap.daum.net"
IMAP_PORT = 993

CHECK_INTERVAL_SECONDS = 10
RECONNECT_DELAY_SECONDS = 10

OPENAI_MODEL = "gpt-4.1-mini"
MAX_BODY_LENGTH = 15_000

# 프로그램 시작 시 기존 메일도 GPT 분석
ANALYZE_EXISTING_MAILS = True

# 확정 주문만 견적서 생성
QUOTE_GENERATION_CATEGORIES = {
    "order"
}

# 업로드한 원본 견적서
TEMPLATE_PATH = Path(
    "25-염치읍사무소.xlsx"
)

# 원본 파일에서 견적서 양식으로 사용할 시트
TEMPLATE_SHEET_NAME = "0619-김주현"

# 견적서 저장 폴더
QUOTE_OUTPUT_DIR = Path(
    "quotation_drafts"
)

QUOTE_OUTPUT_DIR.mkdir(
    parents=True,
    exist_ok=True
)


# =========================================================
# 고정 판매자 정보
# =========================================================

SELLER_NAMES = {
    "열린문디자인",
    "(주)열린문디자인",
    "주식회사 열린문디자인",
}

DEFAULT_DELIVERY_PLACE = "지정장소"
DEFAULT_PAYMENT_TERMS = "현금 또는 카드결제"
DEFAULT_VALIDITY = "견적일로부터"


# =========================================================
# 환경 변수 검증
# =========================================================

if not LOGIN_ID:
    raise RuntimeError(
        ".env 파일에 DAUM_LOGIN_ID를 입력하세요."
    )

if not APP_PASSWORD:
    raise RuntimeError(
        ".env 파일에 DAUM_APP_PASSWORD를 입력하세요."
    )

if not OPENAI_API_KEY:
    raise RuntimeError(
        ".env 파일에 OPENAI_API_KEY를 입력하세요."
    )

if not TEMPLATE_PATH.exists():
    raise FileNotFoundError(
        "견적서 원본 파일을 찾을 수 없습니다.\n"
        f"필요한 위치: {TEMPLATE_PATH.resolve()}"
    )


# =========================================================
# OpenAI 클라이언트
# =========================================================

openai_client = OpenAI(
    api_key=OPENAI_API_KEY
)


# =========================================================
# GPT 구조화 출력 모델
#
# 클래스가 함수보다 먼저 정의되어 있으므로
# MailAnalysis NameError가 발생하지 않는다.
# =========================================================

class OrderItem(BaseModel):
    product_name: str = Field(
        description=(
            "주문한 제품명. "
            "예: 명함, 현수막, 리플릿"
        )
    )

    specification: str | None = Field(
        default=None,
        description=(
            "제품 규격, 크기, 재질, 용지, 색상, "
            "디자인 변경 사항 등의 상세 규격"
        )
    )

    quantity: float | None = Field(
        default=None,
        description=(
            "주문 수량. "
            "메일에서 확인할 수 없으면 null"
        )
    )

    unit: str | None = Field(
        default=None,
        description=(
            "곽, 장, 개, 부, 세트, 매 등의 단위"
        )
    )

    unit_price: int | None = Field(
        default=None,
        description=(
            "메일에 명시된 개당 단가. "
            "메일에 없으면 추측하지 말고 null"
        )
    )

    amount: int | None = Field(
        default=None,
        description=(
            "해당 품목 총 금액. "
            "메일에 명시되어 있거나 "
            "수량과 단가로 확실하게 계산 가능할 때만 입력"
        )
    )

    detail_text: str | None = Field(
        default=None,
        description=(
            "인쇄 문구, 현수막 문구, "
            "디자인 내용 등 품목 아래 별도 행에 들어갈 내용"
        )
    )

    schedule_note: str | None = Field(
        default=None,
        description=(
            "시공, 철거, 방문수령, 배송, 납품 일정 등의 비고"
        )
    )


class MailAnalysis(BaseModel):
    category: Literal[
        "order",
        "quotation_request",
        "advertisement",
        "inquiry",
        "shipping",
        "payment",
        "other",
    ] = Field(
        description="메일의 업무 분류"
    )

    is_order_related: bool = Field(
        description=(
            "주문 또는 견적 업무와 관련되어 있으면 true"
        )
    )

    confidence: float = Field(
        ge=0,
        le=1,
        description="분류 확신도"
    )

    recipient_organization: str | None = Field(
        default=None,
        description=(
            "주문을 보낸 고객 측 회사나 기관명. "
            "열린문디자인은 판매자이므로 제외"
        )
    )

    recipient_department: str | None = Field(
        default=None,
        description=(
            "고객 측 부서나 견적서 수신 담당자. "
            "예: 인사총무팀 담당자 귀하"
        )
    )

    customer_name: str | None = Field(
        default=None,
        description="주문자 또는 고객 담당자 이름"
    )

    customer_phone: str | None = Field(
        default=None,
        description="주문자 전화번호"
    )

    customer_email: str | None = Field(
        default=None,
        description="주문자 이메일 주소"
    )

    delivery_place: str | None = Field(
        default=None,
        description=(
            "구체적인 납품 장소 또는 배송 주소. "
            "'공주 주소'처럼 실제 주소가 아닌 표현이면 원문 그대로 입력"
        )
    )

    payment_terms: str | None = Field(
        default=None,
        description="메일에 명시된 결제 조건"
    )

    requested_date: str | None = Field(
        default=None,
        description="희망 납품일 또는 작업일"
    )

    total_amount: int | None = Field(
        default=None,
        description=(
            "메일에 명시된 전체 주문 금액. "
            "없으면 추측하지 말고 null"
        )
    )

    items: list[OrderItem] = Field(
        default_factory=list,
        description="주문 품목 목록"
    )

    summary: str = Field(
        min_length=1,
        description="메일 핵심 내용 요약"
    )

    reason: str = Field(
        min_length=1,
        description=(
            "해당 분류로 판단한 구체적인 근거를 "
            "메일 원문 표현에 근거해 작성"
        )
    )

    missing_information: list[str] = Field(
        default_factory=list,
        description=(
            "견적서 작성에 필요하지만 메일에 없는 정보"
        )
    )


# =========================================================
# MIME 문자열 디코딩
# =========================================================

def decode_mime_text(
    value: str | None
) -> str:
    if not value:
        return ""

    result: list[str] = []

    for part, charset in decode_header(value):
        if isinstance(part, bytes):
            try:
                result.append(
                    part.decode(
                        charset or "utf-8",
                        errors="replace"
                    )
                )

            except LookupError:
                result.append(
                    part.decode(
                        "utf-8",
                        errors="replace"
                    )
                )

        else:
            result.append(part)

    return "".join(result).strip()


# =========================================================
# 이메일 주소 디코딩
# =========================================================

def decode_email_address(
    value: str | None
) -> str:
    if not value:
        return ""

    name, address = parseaddr(value)
    decoded_name = decode_mime_text(name)

    if decoded_name and address:
        return f"{decoded_name} <{address}>"

    return address or decoded_name


def extract_email_address(
    value: str
) -> str:
    _, address = parseaddr(value)
    return address.strip()


# =========================================================
# 메일 본문 파트 디코딩
# =========================================================

def decode_payload(
    part: Message
) -> str:
    payload = part.get_payload(
        decode=True
    )

    if payload is None:
        return ""

    charset = (
        part.get_content_charset()
        or "utf-8"
    )

    try:
        return payload.decode(
            charset,
            errors="replace"
        )

    except LookupError:
        return payload.decode(
            "utf-8",
            errors="replace"
        )


# =========================================================
# HTML을 텍스트로 변환
# =========================================================

def html_to_text(
    html: str
) -> str:
    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    for tag in soup([
        "script",
        "style",
        "head",
        "meta",
        "link",
        "noscript",
        "svg",
    ]):
        tag.decompose()

    text = soup.get_text(
        separator="\n"
    )

    lines: list[str] = []
    previous_line: str | None = None

    for line in text.splitlines():
        cleaned_line = " ".join(
            line.strip().split()
        )

        if not cleaned_line:
            continue

        if cleaned_line == previous_line:
            continue

        lines.append(cleaned_line)
        previous_line = cleaned_line

    return "\n".join(lines)


# =========================================================
# 메일 본문 추출
# =========================================================

def get_mail_body(
    message: Message
) -> str:
    plain_bodies: list[str] = []
    html_bodies: list[str] = []

    if message.is_multipart():
        for part in message.walk():
            content_type = (
                part.get_content_type()
            )

            disposition = str(
                part.get(
                    "Content-Disposition",
                    ""
                )
            ).lower()

            if part.get_filename():
                continue

            if "attachment" in disposition:
                continue

            if content_type == "text/plain":
                text = decode_payload(
                    part
                ).strip()

                if text:
                    plain_bodies.append(text)

            elif content_type == "text/html":
                html = decode_payload(
                    part
                ).strip()

                if html:
                    html_bodies.append(html)

        if plain_bodies:
            return "\n\n".join(
                plain_bodies
            ).strip()

        if html_bodies:
            return "\n\n".join(
                html_to_text(html)
                for html in html_bodies
            ).strip()

        return ""

    body = decode_payload(message)

    if message.get_content_type() == "text/html":
        return html_to_text(body)

    return body.strip()


# =========================================================
# 첨부파일 이름 목록
# =========================================================

def get_attachment_names(
    message: Message
) -> list[str]:
    names: list[str] = []

    for part in message.walk():
        filename = part.get_filename()

        if not filename:
            continue

        decoded_filename = decode_mime_text(
            filename
        )

        if decoded_filename:
            names.append(decoded_filename)

    return names


# =========================================================
# GPT 메일 분석
# =========================================================

def analyze_mail_with_llm(
    subject: str,
    sender: str,
    receiver: str,
    body: str,
    attachment_names: list[str]
) -> MailAnalysis | None:
    sender_email = extract_email_address(
        sender
    )

    attachment_text = (
        ", ".join(attachment_names)
        if attachment_names
        else "없음"
    )

    prompt = f"""
다음 이메일을 디자인·인쇄 회사의 주문 접수 관점에서 분석하세요.

[판매자 회사]

판매자는 "(주)열린문디자인"입니다.

"열린문디자인", "(주)열린문디자인",
"주식회사 열린문디자인"은 견적서를 작성하는 판매자입니다.

recipient_organization에는 판매자가 아니라
주문 메일을 보낸 고객 측 회사나 기관을 넣어야 합니다.

[분류 기준]

order:
- 고객이 제작, 발주, 주문 진행을 확정적으로 요청
- "제작해주세요", "진행해주세요", "주문합니다",
  "발주합니다" 등의 의사가 확인됨
- 품목, 수량, 규격 등의 주문 정보가 있음

quotation_request:
- 견적, 가격, 단가만 문의
- 아직 제작 또는 주문 진행이 확정되지 않음

advertisement:
- 할인, 이벤트, 뉴스레터, 가입 유도,
  프로모션 또는 불특정 다수 대상 홍보

inquiry:
- 단순 문의이며 주문이나 견적 의사가 명확하지 않음

shipping:
- 배송, 출고, 운송장, 납품 상태 관련

payment:
- 결제, 입금, 세금계산서 관련

other:
- 위 분류에 해당하지 않음

[중요 규칙]

1. 메일에 없는 단가와 금액은 절대 추측하지 마세요.
2. 단가가 없으면 unit_price는 null입니다.
3. 전체 금액이 없으면 total_amount는 null입니다.
4. 제품의 용지, 크기, 재질, 색상, 디자인 변경 내용은
   specification에 넣으세요.
5. 인쇄 문구나 현수막 문구는 detail_text에 넣으세요.
6. 배송, 시공, 철거, 방문수령 일정은 schedule_note에 넣으세요.
7. 발신자 이메일은 customer_email로 사용할 수 있습니다.
8. recipient_organization은 고객 측 기관명입니다.
9. 열린문디자인은 recipient_organization으로 반환하지 마세요.
10. 메일 제목의 [충남연구원] 같은 대괄호 기관명을 우선 확인하세요.
11. 발신자 서명, 부서명, 이메일 도메인도 기관명 판단에 활용하세요.
12. 수신자(To)에 적힌 열린문디자인은 고객 기관이 아닙니다.
13. 확정 주문인 경우에만 category를 order로 분류하세요.
14. 광고는 절대 주문으로 분류하지 마세요.
15. reason은 반드시 비우지 말고 구체적으로 작성하세요.
16. 원문에 없는 정보는 만들어내지 마세요.

[메일]

제목:
{subject}

보낸 사람:
{sender}

발신자 이메일:
{sender_email}

받는 사람:
{receiver}

첨부파일:
{attachment_text}

본문:
{body[:MAX_BODY_LENGTH]}
"""

    try:
        completion = (
            openai_client.beta.chat.completions.parse(
                model=OPENAI_MODEL,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "당신은 디자인·인쇄 업체의 "
                            "주문 메일 분류 및 견적서 초안 작성 AI입니다. "
                            "메일에 없는 정보는 추측하지 않습니다."
                        )
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                response_format=MailAnalysis,
                temperature=0
            )
        )

        result = (
            completion
            .choices[0]
            .message
            .parsed
        )

        if result is None:
            print(
                "GPT 분석 결과가 비어 있습니다."
            )
            return None

        return result

    except Exception as error:
        print()
        print(
            "OpenAI 분석 실패:",
            type(error).__name__,
            error
        )
        return None


# =========================================================
# 고객 기관명 보정
# =========================================================

def resolve_recipient_organization(
    analysis: MailAnalysis,
    subject: str
) -> str:
    organization = (
        analysis.recipient_organization
        or ""
    ).strip()

    normalized = (
        organization
        .replace(" ", "")
        .replace("(주)", "")
        .replace("주식회사", "")
    )

    normalized_seller_names = {
        name
        .replace(" ", "")
        .replace("(주)", "")
        .replace("주식회사", "")
        for name in SELLER_NAMES
    }

    if (
        organization
        and normalized not in normalized_seller_names
    ):
        return organization

    # 제목의 [충남연구원] 같은 기관명 추출
    match = re.search(
        r"\[([^\]]+)\]",
        subject
    )

    if match:
        candidate = match.group(1).strip()

        candidate_normalized = (
            candidate
            .replace(" ", "")
            .replace("(주)", "")
            .replace("주식회사", "")
        )

        if (
            candidate_normalized
            not in normalized_seller_names
        ):
            return candidate

    return (
        analysis.customer_name
        or "담당 기관"
    )


# =========================================================
# 숫자 금액을 한글로 변환
# =========================================================

def number_to_korean_won(
    number: int
) -> str:
    if number <= 0:
        return ""

    digit_names = [
        "",
        "일",
        "이",
        "삼",
        "사",
        "오",
        "육",
        "칠",
        "팔",
        "구",
    ]

    small_units = [
        "",
        "십",
        "백",
        "천",
    ]

    large_units = [
        "",
        "만",
        "억",
        "조",
    ]

    groups: list[int] = []
    remaining = number

    while remaining > 0:
        groups.append(
            remaining % 10_000
        )
        remaining //= 10_000

    result_parts: list[str] = []

    for group_index in range(
        len(groups) - 1,
        -1,
        -1
    ):
        group = groups[group_index]

        if group == 0:
            continue

        group_text = ""

        for position in range(
            3,
            -1,
            -1
        ):
            divisor = 10 ** position
            digit = (
                group // divisor
            ) % 10

            if digit == 0:
                continue

            if digit == 1 and position > 0:
                group_text += (
                    small_units[position]
                )

            else:
                group_text += (
                    digit_names[digit]
                    + small_units[position]
                )

        group_text += (
            large_units[group_index]
        )

        result_parts.append(
            group_text
        )

    return "".join(result_parts)


# =========================================================
# 파일명 정리
# =========================================================

def sanitize_filename(
    value: str,
    max_length: int = 30
) -> str:
    if not value:
        return "미확인"

    cleaned = re.sub(
        r'[\\/:*?"<>|]',
        "_",
        value
    )

    cleaned = re.sub(
        r"\s+",
        " ",
        cleaned
    ).strip()

    cleaned = cleaned.rstrip(
        ". "
    )

    if not cleaned:
        return "미확인"

    return cleaned[:max_length]


# =========================================================
# 품목 수량 표시
# =========================================================

def format_quantity(
    quantity: float | None
) -> int | float | None:
    if quantity is None:
        return None

    numeric_quantity = float(
        quantity
    )

    if numeric_quantity.is_integer():
        return int(
            numeric_quantity
        )

    return numeric_quantity


# =========================================================
# 품목 및 규격 문자열 생성
# =========================================================

def build_item_text(
    item: OrderItem
) -> str:
    lines: list[str] = []

    if item.product_name:
        lines.append(
            item.product_name
        )

    specification = (
        item.specification
        or ""
    ).strip()

    if specification:
        lines.append(
            f"({specification})"
        )

    return "\n".join(lines)


# =========================================================
# 총액 계산
# =========================================================

def calculate_total_amount(
    analysis: MailAnalysis
) -> int | None:
    if analysis.total_amount is not None:
        return int(
            analysis.total_amount
        )

    total = 0
    has_calculable_amount = False

    for item in analysis.items:
        if item.amount is not None:
            total += int(
                item.amount
            )
            has_calculable_amount = True

        elif (
            item.quantity is not None
            and item.unit_price is not None
        ):
            total += int(
                item.quantity
                * item.unit_price
            )
            has_calculable_amount = True

    if not has_calculable_amount:
        return None

    return total


# =========================================================
# 안전한 출력 파일 경로 생성
# =========================================================

def make_output_path(
    uid: bytes,
    recipient: str,
    subject: str
) -> Path:
    recipient_filename = sanitize_filename(
        recipient
    )

    subject_filename = sanitize_filename(
        subject
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S_%f"
    )

    filename = (
        f"견적서초안_UID{uid.decode()}_"
        f"{recipient_filename}_"
        f"{subject_filename}_"
        f"{timestamp}.xlsx"
    )

    return QUOTE_OUTPUT_DIR / filename


# =========================================================
# 원본 템플릿 기반 견적서 생성
#
# 업로드한 원본 실제 셀 배치:
#
# B3       수신 기관
# D4       수신 담당자
# D5       견적일
# D6       납품장소
# D7       결제조건
# D8       유효기간
# L5       주문자 이름
# L6       주문자 번호
# L7       주문자 이메일
# D10      한글 공급금액
# I10      숫자 공급금액
# C14:C23  품목 및 규격
# F14:F23  수량
# G14:G23  단가
# I14:I23  공급금액
# L14:L23  비고
# G24      하단 공급금액
# =========================================================

def create_quotation_excel(
    uid: bytes,
    subject: str,
    analysis: MailAnalysis
) -> Path:
    recipient = resolve_recipient_organization(
        analysis=analysis,
        subject=subject
    )

    output_path = make_output_path(
        uid=uid,
        recipient=recipient,
        subject=subject
    )

    # 원본 파일 자체를 먼저 복사한다.
    # 이미지, 병합, 행 높이, 열 너비, 테두리를 그대로 유지한다.
    shutil.copy2(
        TEMPLATE_PATH,
        output_path
    )

    workbook = load_workbook(
        output_path
    )

    if (
        TEMPLATE_SHEET_NAME
        not in workbook.sheetnames
    ):
        output_path.unlink(
            missing_ok=True
        )

        raise KeyError(
            "지정한 견적서 템플릿 시트를 찾을 수 없습니다.\n"
            f"요청 시트: {TEMPLATE_SHEET_NAME}\n"
            f"실제 시트: {workbook.sheetnames}"
        )

    sheet = workbook[
        TEMPLATE_SHEET_NAME
    ]

    # 대상 템플릿 시트만 남긴다.
    # 시트 자체를 새로 만들지 않으므로 해당 시트의 서식과 이미지 유지.
    for worksheet in list(
        workbook.worksheets
    ):
        if worksheet.title != TEMPLATE_SHEET_NAME:
            workbook.remove(
                worksheet
            )

    sheet.title = "견적서"
    workbook.active = 0

    # -----------------------------------------------------
    # 기존 고객 및 견적 데이터 초기화
    # -----------------------------------------------------

    sheet["B3"] = None
    sheet["D4"] = None
    sheet["D5"] = None
    sheet["D6"] = None
    sheet["D7"] = None
    sheet["D8"] = None

    sheet["L5"] = None
    sheet["L6"] = None
    sheet["L7"] = None

    # 품목 1~10 초기화
    for row in range(14, 24):
        sheet[f"C{row}"] = None
        sheet[f"F{row}"] = None
        sheet[f"G{row}"] = None
        sheet[f"I{row}"] = None
        sheet[f"L{row}"] = None

    # -----------------------------------------------------
    # 견적서 상단 입력
    # -----------------------------------------------------

    sheet["B3"] = (
        f"{recipient} 귀하"
    )

    sheet["D4"] = (
        analysis.recipient_department
        or "담당자 귀하"
    )

    # 원본 셀의 날짜 서식을 유지하도록 날짜 객체 입력
    sheet["D5"] = datetime.now()

    sheet["D6"] = (
        analysis.delivery_place
        or DEFAULT_DELIVERY_PLACE
    )

    sheet["D7"] = (
        analysis.payment_terms
        or DEFAULT_PAYMENT_TERMS
    )

    sheet["D8"] = DEFAULT_VALIDITY

    sheet["L5"] = (
        analysis.customer_name
        or ""
    )

    sheet["L6"] = (
        analysis.customer_phone
        or ""
    )

    sheet["L7"] = (
        analysis.customer_email
        or ""
    )

    # -----------------------------------------------------
    # 품목 입력
    # -----------------------------------------------------

    current_row = 14

    for item in analysis.items:
        if current_row > 23:
            break

        sheet[f"C{current_row}"] = (
            build_item_text(item)
        )

        sheet[f"F{current_row}"] = (
            format_quantity(
                item.quantity
            )
        )

        if item.unit_price is not None:
            sheet[f"G{current_row}"] = int(
                item.unit_price
            )

        if item.amount is not None:
            sheet[f"I{current_row}"] = int(
                item.amount
            )

        elif (
            item.quantity is not None
            and item.unit_price is not None
        ):
            sheet[f"I{current_row}"] = (
                f"=F{current_row}*G{current_row}"
            )

        if item.schedule_note:
            sheet[f"L{current_row}"] = (
                item.schedule_note
            )

        # 현수막 문구 등은 다음 행에 따로 배치
        if (
            item.detail_text
            and current_row < 23
        ):
            detail_row = (
                current_row + 1
            )

            sheet[f"C{detail_row}"] = (
                item.detail_text
            )

            current_row += 2

        else:
            current_row += 1

    # -----------------------------------------------------
    # 공급금액
    # -----------------------------------------------------

    total_amount = calculate_total_amount(
        analysis
    )

    if total_amount is not None:
        # 원본 파일의 수식 구조 유지
        # D10은 하단 총액을 참조
        # I10은 D10을 참조
        # G24는 품목 합계를 계산
        sheet["G24"] = "=SUM(I14:K23)"
        sheet["D10"] = "=G24"
        sheet["I10"] = "=D10"

        # Excel에서 열었을 때 수식 재계산
        try:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
        except AttributeError:
            pass

    else:
        # 금액을 알 수 없으면 0원이 아니라 빈칸
        sheet["D10"] = ""
        sheet["I10"] = ""
        sheet["G24"] = ""

    # -----------------------------------------------------
    # 저장
    # -----------------------------------------------------

    try:
        workbook.save(
            output_path
        )

    except PermissionError as error:
        try:
            workbook.close()
        except Exception:
            pass

        raise PermissionError(
            "견적서 파일을 저장하지 못했습니다.\n"
            "같은 파일이 Excel에서 열려 있는지 확인하세요.\n"
            f"저장 경로: {output_path.resolve()}"
        ) from error

    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return output_path


# =========================================================
# GPT 분석 결과 출력
# =========================================================

def print_analysis(
    analysis: MailAnalysis,
    subject: str
) -> None:
    category_labels = {
        "order": "주문",
        "quotation_request": "견적 요청",
        "advertisement": "광고",
        "inquiry": "일반 문의",
        "shipping": "배송",
        "payment": "결제",
        "other": "기타",
    }

    recipient = resolve_recipient_organization(
        analysis=analysis,
        subject=subject
    )

    print()
    print("[GPT 분석 결과]")

    print(
        "분류:",
        category_labels.get(
            analysis.category,
            analysis.category
        )
    )

    print(
        "주문 관련:",
        "예"
        if analysis.is_order_related
        else "아니오"
    )

    print(
        "확신도:",
        f"{analysis.confidence:.2f}"
    )

    print(
        "요약:",
        analysis.summary
    )

    print(
        "판단 근거:",
        analysis.reason
    )

    print(
        "수신 기관:",
        recipient
    )

    print(
        "수신 담당:",
        analysis.recipient_department
        or "미확인"
    )

    print(
        "주문자:",
        analysis.customer_name
        or "미확인"
    )

    print(
        "전화번호:",
        analysis.customer_phone
        or "미확인"
    )

    print(
        "이메일:",
        analysis.customer_email
        or "미확인"
    )

    print(
        "납품장소:",
        analysis.delivery_place
        or "미확인"
    )

    print(
        "전체 금액:",
        (
            f"{analysis.total_amount:,}원"
            if analysis.total_amount is not None
            else "미확인"
        )
    )

    if analysis.items:
        print("품목:")

        for index, item in enumerate(
            analysis.items,
            start=1
        ):
            quantity_text = (
                format_quantity(
                    item.quantity
                )
                if item.quantity is not None
                else "미확인"
            )

            print(
                f"  {index}. {item.product_name}"
            )

            print(
                "     규격:",
                item.specification
                or "미확인"
            )

            print(
                "     수량:",
                quantity_text,
                item.unit or ""
            )

            print(
                "     단가:",
                (
                    f"{item.unit_price:,}원"
                    if item.unit_price is not None
                    else "미확인"
                )
            )

            print(
                "     금액:",
                (
                    f"{item.amount:,}원"
                    if item.amount is not None
                    else "미확인"
                )
            )

            if item.detail_text:
                print(
                    "     상세 문구:",
                    item.detail_text
                )

            if item.schedule_note:
                print(
                    "     일정:",
                    item.schedule_note
                )

    else:
        print("품목: 없음")

    if analysis.missing_information:
        print(
            "누락 정보:",
            ", ".join(
                analysis.missing_information
            )
        )

    else:
        print("누락 정보: 없음")


# =========================================================
# IMAP 연결
# =========================================================

def connect_imap() -> imaplib.IMAP4_SSL:
    print()
    print(
        "Daum IMAP 서버 연결 중..."
    )

    imap = imaplib.IMAP4_SSL(
        IMAP_SERVER,
        IMAP_PORT
    )

    imap.login(
        LOGIN_ID,
        APP_PASSWORD
    )

    status, data = imap.select(
        "INBOX",
        readonly=True
    )

    if status != "OK":
        raise RuntimeError(
            "받은편지함을 열 수 없습니다."
        )

    mail_count = (
        data[0].decode()
        if data and data[0]
        else "0"
    )

    print("로그인 성공")

    print(
        "현재 받은편지함 메일 수:",
        mail_count
    )

    return imap


# =========================================================
# UID 조회
# =========================================================

def get_all_uids(
    imap: imaplib.IMAP4_SSL
) -> list[bytes]:
    status, data = imap.uid(
        "search",
        None,
        "ALL"
    )

    if status != "OK":
        raise RuntimeError(
            "메일 UID 검색에 실패했습니다."
        )

    if not data or not data[0]:
        return []

    return data[0].split()


def get_new_uids(
    imap: imaplib.IMAP4_SSL,
    last_uid: int
) -> list[bytes]:
    status, data = imap.uid(
        "search",
        None,
        f"UID {last_uid + 1}:*"
    )

    if status != "OK":
        raise RuntimeError(
            "새 메일 UID 검색에 실패했습니다."
        )

    if not data or not data[0]:
        return []

    result: list[bytes] = []

    for uid in data[0].split():
        try:
            uid_number = int(uid)

            if uid_number > last_uid:
                result.append(uid)

        except ValueError:
            continue

    return result


# =========================================================
# 특정 메일 가져오기
# =========================================================

def fetch_mail(
    imap: imaplib.IMAP4_SSL,
    uid: bytes
) -> Message | None:
    status, data = imap.uid(
        "fetch",
        uid,
        "(BODY.PEEK[])"
    )

    if status != "OK":
        print(
            "메일 가져오기 실패 UID:",
            uid.decode()
        )
        return None

    for response in data:
        if not isinstance(
            response,
            tuple
        ):
            continue

        raw_mail = response[1]

        if raw_mail:
            return email.message_from_bytes(
                raw_mail
            )

    return None


# =========================================================
# 개별 메일 처리
# =========================================================

def process_mail(
    uid: bytes,
    message: Message,
    mail_label: str
) -> None:
    subject = decode_mime_text(
        message.get("Subject")
    )

    sender = decode_email_address(
        message.get("From")
    )

    receiver = decode_email_address(
        message.get("To")
    )

    cc = decode_email_address(
        message.get("Cc")
    )

    date = decode_mime_text(
        message.get("Date")
    )

    message_id = decode_mime_text(
        message.get("Message-ID")
    )

    body = get_mail_body(
        message
    )

    attachment_names = (
        get_attachment_names(
            message
        )
    )

    print()
    print("=" * 100)
    print(mail_label)
    print("UID:", uid.decode())
    print("Message-ID:", message_id)
    print("제목:", subject)
    print("보낸 사람:", sender)
    print("받는 사람:", receiver)

    if cc:
        print("참조:", cc)

    print("날짜:", date)

    print(
        "첨부파일:",
        (
            ", ".join(
                attachment_names
            )
            if attachment_names
            else "없음"
        )
    )

    print("-" * 100)
    print("본문:")
    print(body or "[본문 없음]")

    print("-" * 100)
    print("GPT 주문 여부 분석 중...")

    analysis = analyze_mail_with_llm(
        subject=subject,
        sender=sender,
        receiver=receiver,
        body=body,
        attachment_names=attachment_names
    )

    if analysis is None:
        print("GPT 분석 실패")
        print("=" * 100)
        return

    print_analysis(
        analysis=analysis,
        subject=subject
    )

    # 확정 주문인 경우에만 견적서 생성
    if (
        analysis.category
        in QUOTE_GENERATION_CATEGORIES
        and analysis.is_order_related
    ):
        try:
            output_path = (
                create_quotation_excel(
                    uid=uid,
                    subject=subject,
                    analysis=analysis
                )
            )

            print()
            print(
                "견적서 초안 생성 완료:"
            )

            print(
                output_path.resolve()
            )

        except Exception as error:
            print()
            print(
                "견적서 생성 실패:",
                type(error).__name__,
                error
            )

    else:
        print()
        print(
            "확정 주문 메일이 아니므로 "
            "견적서를 생성하지 않습니다."
        )

    print("=" * 100)


# =========================================================
# 메일 감시
# =========================================================

def monitor_new_mails() -> None:
    last_uid = 0
    initial_load_completed = False

    print(
        "Daum 주문 메일 감시 및 "
        "견적서 생성 프로그램 시작"
    )

    print(
        "사용 모델:",
        OPENAI_MODEL
    )

    print(
        "견적서 템플릿:",
        TEMPLATE_PATH.resolve()
    )

    print(
        "템플릿 시트:",
        TEMPLATE_SHEET_NAME
    )

    print(
        f"{CHECK_INTERVAL_SECONDS}초마다 "
        "새 메일을 확인합니다."
    )

    print(
        "종료하려면 Ctrl + C를 누르세요."
    )

    while True:
        imap: imaplib.IMAP4_SSL | None = None

        try:
            imap = connect_imap()

            # -------------------------------------------------
            # 최초 실행 시 기존 메일 처리
            # -------------------------------------------------

            if not initial_load_completed:
                all_uids = get_all_uids(
                    imap
                )

                print()
                print(
                    f"기존 메일 "
                    f"{len(all_uids)}개를 확인합니다."
                )

                if ANALYZE_EXISTING_MAILS:
                    for index, uid in enumerate(
                        all_uids,
                        start=1
                    ):
                        message = fetch_mail(
                            imap,
                            uid
                        )

                        if message is None:
                            continue

                        process_mail(
                            uid=uid,
                            message=message,
                            mail_label=(
                                f"기존 메일 "
                                f"[{index}/{len(all_uids)}]"
                            )
                        )

                if all_uids:
                    last_uid = int(
                        all_uids[-1]
                    )

                else:
                    last_uid = 0

                initial_load_completed = True

                print()
                print(
                    "기존 메일 처리 완료"
                )

                print(
                    "현재 마지막 UID:",
                    last_uid
                )

                print(
                    "이제부터 새 메일을 감시합니다."
                )

            # -------------------------------------------------
            # 새 메일 반복 감시
            # -------------------------------------------------

            while True:
                status, _ = imap.noop()

                if status != "OK":
                    raise imaplib.IMAP4.abort(
                        "NOOP 응답 실패"
                    )

                new_uids = get_new_uids(
                    imap=imap,
                    last_uid=last_uid
                )

                if new_uids:
                    print()
                    print(
                        f"새 메일 "
                        f"{len(new_uids)}개 감지"
                    )

                for uid in new_uids:
                    message = fetch_mail(
                        imap,
                        uid
                    )

                    if message is not None:
                        process_mail(
                            uid=uid,
                            message=message,
                            mail_label="새 메일 도착"
                        )

                    uid_number = int(uid)

                    if uid_number > last_uid:
                        last_uid = uid_number

                time.sleep(
                    CHECK_INTERVAL_SECONDS
                )

        except KeyboardInterrupt:
            print()
            print(
                "메일 감시를 종료합니다."
            )
            break

        except imaplib.IMAP4.error as error:
            print()
            print(
                "IMAP 오류:",
                error
            )

            print(
                f"{RECONNECT_DELAY_SECONDS}초 후 "
                "다시 연결합니다."
            )

            time.sleep(
                RECONNECT_DELAY_SECONDS
            )

        except (
            OSError,
            TimeoutError,
            ConnectionError
        ) as error:
            print()
            print(
                "네트워크 오류:",
                error
            )

            print(
                f"{RECONNECT_DELAY_SECONDS}초 후 "
                "다시 연결합니다."
            )

            time.sleep(
                RECONNECT_DELAY_SECONDS
            )

        except Exception as error:
            print()
            print(
                "예상하지 못한 오류:",
                type(error).__name__,
                error
            )

            print(
                f"{RECONNECT_DELAY_SECONDS}초 후 "
                "다시 연결합니다."
            )

            time.sleep(
                RECONNECT_DELAY_SECONDS
            )

        finally:
            if imap is not None:
                try:
                    imap.logout()
                except Exception:
                    pass


# =========================================================
# 실행
# =========================================================

if __name__ == "__main__":
    monitor_new_mails()