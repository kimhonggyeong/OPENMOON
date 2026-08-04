from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import traceback

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================================================
# 기본 설정
# =========================================================

DEFAULT_INPUT_PATH = Path(
    "26년_열린문디자인_단가표_8차(260528).xlsx"
)

DEFAULT_DB_PATH = Path(
    "price_table.db"
)

SUPPORTED_PRICE_HEADERS = {
    "단가",
    "판매가",
    "판매단가",
    "공급가",
    "공급금액",
    "금액",
}

BLOCKED_ITEM_WORDS = {
    "단가",
    "원가",
    "마진",
    "금액",
    "공급금액",
    "수량",
    "구분",
    "규격",
    "사이즈",
    "세로",
    "가로",
    "비고",
    "업체",
    "업체명",
    "부가세",
    "부가세포함",
    "부가세별도",
    "참고사항",
    "종류",
    "두께",
    "재질",
    "원",
    "천원",
}

NOTE_PREFIXES = (
    "*",
    "★",
    "※",
    "-",
    "ㆍ",
)

SIZE_PATTERN = re.compile(
    r"(?<!\d)"
    r"(\d+(?:\.\d+)?)"
    r"\s*(mm|㎜|cm|㎝|m|인치|inch|in)?"
    r"\s*[*xX×ｘ]\s*"
    r"(\d+(?:\.\d+)?)"
    r"\s*(mm|㎜|cm|㎝|m|인치|inch|in)?",
    re.IGNORECASE,
)

RANGE_PATTERN = re.compile(
    r"^\s*(\d+(?:\.\d+)?)\s*[~\-]\s*(\d+(?:\.\d+)?)\s*$"
)

THICKNESS_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s*T\b",
    re.IGNORECASE,
)

QUANTITY_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s*(매|부|장|개|곽|권|세트|조|롤|박스|식)?"
)

PAPER_PATTERN = re.compile(
    r"([가-힣A-Za-z0-9\-]+)\s*(\d{2,4})\s*g",
    re.IGNORECASE,
)


# =========================================================
# 품목 표준화
# =========================================================

NORMALIZED_ALIASES = {
    "현수막": (
        "현수막",
        "게릴라",
        "게시대",
        "육교현수막",
        "시청현수막",
    ),
    "배너": (
        "배너",
        "미니배너",
        "롤업배너",
        "패트지",
        "메쉬",
        "부직포",
    ),
    "배너대": (
        "배너대",
    ),
    "어깨띠": (
        "어깨띠",
    ),
    "사원증": (
        "사원증",
        "명찰",
        "id카드",
    ),
    "인포그래픽": (
        "인포그래픽",
    ),
    "책제본": (
        "책제본",
        "제본",
    ),
    "명함": (
        "명함",
        "카드명함",
        "점자명함",
    ),
    "전단지": (
        "전단지",
        "전단",
    ),
    "포스터": (
        "포스터",
    ),
    "리플릿": (
        "리플릿",
        "리플렛",
    ),
    "카다로그": (
        "카다로그",
        "카탈로그",
        "브로슈어",
        "브로셔",
    ),
    "봉투": (
        "봉투",
    ),
    "상장지": (
        "상장지",
        "상장",
    ),
    "양식지": (
        "양식지",
    ),
    "골지보드": (
        "골지",
        "허니콤",
        "보드",
    ),
    "포맥스": (
        "포맥스",
    ),
    "아크릴": (
        "아크릴",
    ),
    "친환경배너": (
        "친환경배너",
        "친환경 배너",
    ),
}


# =========================================================
# 데이터 구조
# =========================================================

@dataclass
class PriceItem:
    product_name: str
    normalized_name: str
    category: str | None = None
    specification: str | None = None

    width_mm: float | None = None
    height_mm: float | None = None
    width_mm_min: float | None = None
    width_mm_max: float | None = None
    height_mm_min: float | None = None
    height_mm_max: float | None = None

    thickness_mm: float | None = None
    material: str | None = None
    paper: str | None = None
    color: str | None = None
    print_side: str | None = None

    quantity: float | None = None
    quantity_min: float | None = None
    quantity_max: float | None = None
    unit: str | None = None

    unit_price: int | None = None
    total_price: int | None = None
    vat_included: int | None = None

    sheet_name: str = ""
    row_number: int | None = None
    column_number: int | None = None
    original_text: str | None = None

    confidence: float = 1.0
    review_required: int = 0


# =========================================================
# 공통 유틸리티
# =========================================================

def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def compact_text(value: Any) -> str:
    return re.sub(
        r"[\s:：·ㆍ\-_()/\[\]{}]",
        "",
        normalize_text(value)
    ).lower()


def is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False

    return isinstance(value, (int, float))


def to_int_price(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return int(round(value))

    text = normalize_text(value)
    if not text or text.startswith("="):
        return None

    match = re.fullmatch(
        r"\s*₩?\s*([\d,]+(?:\.\d+)?)\s*(?:원)?\s*",
        text
    )

    if not match:
        return None

    try:
        number = float(
            match.group(1).replace(",", "")
        )
    except ValueError:
        return None

    if number <= 0:
        return None

    return int(round(number))


def convert_to_mm(
    value: float,
    unit: str | None,
    default_unit: str = "mm"
) -> float:
    resolved = (unit or default_unit).lower()

    if resolved in {"cm", "㎝"}:
        return value * 10

    if resolved == "m":
        return value * 1000

    if resolved in {
        "인치",
        "inch",
        "in",
    }:
        return value * 25.4

    return value


def extract_size(
    text: str | None,
    default_unit: str = "mm"
) -> tuple[float | None, float | None]:
    if not text:
        return None, None

    match = SIZE_PATTERN.search(text)
    if not match:
        return None, None

    first = float(match.group(1))
    first_unit = match.group(2)
    second = float(match.group(3))
    second_unit = match.group(4)

    resolved = (
        second_unit
        or first_unit
        or default_unit
    )

    return (
        convert_to_mm(first, resolved, default_unit),
        convert_to_mm(second, resolved, default_unit),
    )


def extract_thickness(text: str | None) -> float | None:
    if not text:
        return None

    match = THICKNESS_PATTERN.search(text)
    if not match:
        return None

    return float(match.group(1))


def parse_range(
    value: Any
) -> tuple[float | None, float | None]:
    text = normalize_text(value)
    if not text:
        return None, None

    match = RANGE_PATTERN.match(text)
    if not match:
        number = to_float(value)
        if number is None:
            return None, None
        return number, number

    return (
        float(match.group(1)),
        float(match.group(2)),
    )


def to_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_text(value)
    if not text:
        return None

    match = re.fullmatch(
        r"\s*([\d,]+(?:\.\d+)?)\s*",
        text
    )

    if not match:
        return None

    try:
        return float(
            match.group(1).replace(",", "")
        )
    except ValueError:
        return None


def extract_quantity(
    value: Any
) -> tuple[float | None, str | None]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value), None

    text = normalize_text(value)
    if not text:
        return None, None

    match = QUANTITY_PATTERN.search(text)
    if not match:
        return None, None

    return (
        float(match.group(1)),
        match.group(2),
    )

def extract_quantity_strict(
    value: Any
) -> tuple[float | None, str | None]:
    """
    범용 파서용 수량 추출.

    숫자만 있는 셀은 수량으로 인정하지 않는다.

    인정:
    100매
    500부
    10장
    2개

    제외:
    90
    150
    250
    -> 종이 평량이나 가격 일부일 수 있음
    """

    text = normalize_text(value)

    if not text:
        return None, None

    match = re.fullmatch(
        r"\s*"
        r"([\d,]+(?:\.\d+)?)"
        r"\s*"
        r"(매|부|장|개|곽|권|세트|조|롤|박스|식)"
        r"\s*",
        text
    )

    if not match:
        return None, None

    quantity = float(
        match.group(1).replace(",", "")
    )

    unit = match.group(2)

    return quantity, unit

def extract_paper(text: str | None) -> str | None:
    if not text:
        return None

    match = PAPER_PATTERN.search(text)
    if not match:
        return None

    return f"{match.group(1)} {match.group(2)}g"


def infer_print_side(text: str | None) -> str | None:
    compact = compact_text(text)

    if "양면" in compact:
        return "양면"

    if "단면" in compact:
        return "단면"

    return None


def normalize_product_name(
    product_name: str,
    sheet_name: str
) -> str:
    haystack = compact_text(
        f"{product_name} {sheet_name}"
    )

    for normalized, aliases in NORMALIZED_ALIASES.items():
        for alias in aliases:
            if compact_text(alias) in haystack:
                return normalized

    return normalize_text(product_name)[:100]


def is_blocked_product_text(text: str) -> bool:
    normalized = normalize_text(text)
    compact = compact_text(normalized)

    if not normalized:
        return True

    if len(normalized) > 100:
        return True

    if normalized.startswith(NOTE_PREFIXES):
        return True

    if compact in {
        compact_text(word)
        for word in BLOCKED_ITEM_WORDS
    }:
        return True

    if "부가세포함" in compact:
        return True

    if "부가세별도" in compact:
        return True

    if "단위" in compact:
        return True

    if "참고사항" in compact:
        return True

    if "제작업체" in compact:
        return True

    return False


def row_original_text(
    ws: Worksheet,
    row: int,
    min_col: int = 1,
    max_col: int | None = None
) -> str:
    if max_col is None:
        max_col = ws.max_column

    values: list[str] = []

    for col in range(min_col, max_col + 1):
        text = normalize_text(
            ws.cell(row, col).value
        )

        if text:
            values.append(
                f"{ws.cell(row, col).coordinate}={text}"
            )

    return " | ".join(values)


def nearest_text_above(
    ws: Worksheet,
    row: int,
    col: int,
    max_distance: int = 5
) -> str | None:
    for target_row in range(
        row - 1,
        max(0, row - max_distance - 1),
        -1
    ):
        text = normalize_text(
            ws.cell(target_row, col).value
        )

        if text:
            return text

    return None


def nearest_text_left(
    ws: Worksheet,
    row: int,
    col: int,
    max_distance: int = 8
) -> str | None:
    for target_col in range(
        col - 1,
        max(0, col - max_distance - 1),
        -1
    ):
        text = normalize_text(
            ws.cell(row, target_col).value
        )

        if text:
            return text

    return None


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()

    with path.open("rb") as file:
        while True:
            chunk = file.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)

    return digest.hexdigest()


# =========================================================
# DB
# =========================================================

def connect_database(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row

    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA journal_mode = WAL")
    connection.execute("PRAGMA synchronous = NORMAL")

    return connection


def create_schema(
    connection: sqlite3.Connection
) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS source_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path TEXT NOT NULL UNIQUE,
            file_name TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            indexed_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS price_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            product_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            category TEXT,
            specification TEXT,

            width_mm REAL,
            height_mm REAL,
            width_mm_min REAL,
            width_mm_max REAL,
            height_mm_min REAL,
            height_mm_max REAL,

            thickness_mm REAL,
            material TEXT,
            paper TEXT,
            color TEXT,
            print_side TEXT,

            quantity REAL,
            quantity_min REAL,
            quantity_max REAL,
            unit TEXT,

            unit_price INTEGER,
            total_price INTEGER,
            vat_included INTEGER,

            sheet_name TEXT NOT NULL,
            row_number INTEGER,
            column_number INTEGER,
            original_text TEXT,

            confidence REAL NOT NULL DEFAULT 1.0,
            review_required INTEGER NOT NULL DEFAULT 0,

            CHECK (
                unit_price IS NOT NULL
                OR total_price IS NOT NULL
            )
        );

        CREATE TABLE IF NOT EXISTS price_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT,
            normalized_name TEXT,
            rule_type TEXT NOT NULL,
            rule_text TEXT NOT NULL,
            numeric_value REAL,
            unit TEXT,
            sheet_name TEXT NOT NULL,
            row_number INTEGER,
            original_text TEXT
        );

        CREATE TABLE IF NOT EXISTS price_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT,
            normalized_name TEXT,
            note_type TEXT NOT NULL,
            note_text TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            row_number INTEGER,
            original_text TEXT
        );

        CREATE TABLE IF NOT EXISTS review_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_name TEXT NOT NULL,
            row_number INTEGER,
            column_number INTEGER,
            reason TEXT NOT NULL,
            candidate_text TEXT,
            candidate_price INTEGER,
            original_text TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_price_items_name
        ON price_items(normalized_name);

        CREATE INDEX IF NOT EXISTS idx_price_items_dimensions
        ON price_items(
            normalized_name,
            width_mm,
            height_mm,
            width_mm_min,
            width_mm_max,
            height_mm_min,
            height_mm_max
        );

        CREATE INDEX IF NOT EXISTS idx_price_items_quantity
        ON price_items(
            normalized_name,
            quantity,
            quantity_min,
            quantity_max
        );

        CREATE INDEX IF NOT EXISTS idx_price_items_material
        ON price_items(
            normalized_name,
            material,
            paper,
            thickness_mm,
            print_side
        );

        CREATE INDEX IF NOT EXISTS idx_price_items_source
        ON price_items(sheet_name, row_number);

        CREATE INDEX IF NOT EXISTS idx_price_rules_name
        ON price_rules(normalized_name);

        CREATE INDEX IF NOT EXISTS idx_price_notes_name
        ON price_notes(normalized_name);
        """
    )

    connection.commit()


def insert_price_item(
    connection: sqlite3.Connection,
    item: PriceItem
) -> None:
    if (
        item.unit_price is None
        and item.total_price is None
    ):
        return

    if is_blocked_product_text(
        item.product_name
    ):
        connection.execute(
            """
            INSERT INTO review_items (
                sheet_name,
                row_number,
                column_number,
                reason,
                candidate_text,
                candidate_price,
                original_text
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.sheet_name,
                item.row_number,
                item.column_number,
                "품목명이 헤더·설명·단위로 판단됨",
                item.product_name,
                item.unit_price or item.total_price,
                item.original_text,
            )
        )
        return

    condition_exists = any(
        value is not None and value != ""
        for value in (
            item.specification,
            item.width_mm,
            item.height_mm,
            item.width_mm_min,
            item.width_mm_max,
            item.height_mm_min,
            item.height_mm_max,
            item.thickness_mm,
            item.material,
            item.paper,
            item.color,
            item.print_side,
            item.quantity,
            item.quantity_min,
            item.quantity_max,
        )
    )

    if not condition_exists:
        connection.execute(
            """
            INSERT INTO review_items (
                sheet_name,
                row_number,
                column_number,
                reason,
                candidate_text,
                candidate_price,
                original_text
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.sheet_name,
                item.row_number,
                item.column_number,
                "가격은 있으나 구매 조건이 없음",
                item.product_name,
                item.unit_price or item.total_price,
                item.original_text,
            )
        )
        return

    connection.execute(
        """
        INSERT INTO price_items (
            product_name,
            normalized_name,
            category,
            specification,

            width_mm,
            height_mm,
            width_mm_min,
            width_mm_max,
            height_mm_min,
            height_mm_max,

            thickness_mm,
            material,
            paper,
            color,
            print_side,

            quantity,
            quantity_min,
            quantity_max,
            unit,

            unit_price,
            total_price,
            vat_included,

            sheet_name,
            row_number,
            column_number,
            original_text,

            confidence,
            review_required
        )
        VALUES (
            ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?
        )
        """,
        (
            item.product_name,
            item.normalized_name,
            item.category,
            item.specification,

            item.width_mm,
            item.height_mm,
            item.width_mm_min,
            item.width_mm_max,
            item.height_mm_min,
            item.height_mm_max,

            item.thickness_mm,
            item.material,
            item.paper,
            item.color,
            item.print_side,

            item.quantity,
            item.quantity_min,
            item.quantity_max,
            item.unit,

            item.unit_price,
            item.total_price,
            item.vat_included,

            item.sheet_name,
            item.row_number,
            item.column_number,
            item.original_text,

            item.confidence,
            item.review_required,
        )
    )


def insert_note(
    connection: sqlite3.Connection,
    sheet_name: str,
    row_number: int,
    text: str,
    note_type: str = "note",
    product_name: str | None = None
) -> None:
    normalized_name = (
        normalize_product_name(
            product_name,
            sheet_name
        )
        if product_name
        else normalize_product_name(
            sheet_name,
            sheet_name
        )
    )

    connection.execute(
        """
        INSERT INTO price_notes (
            product_name,
            normalized_name,
            note_type,
            note_text,
            sheet_name,
            row_number,
            original_text
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            product_name,
            normalized_name,
            note_type,
            text,
            sheet_name,
            row_number,
            text,
        )
    )


def insert_rule(
    connection: sqlite3.Connection,
    sheet_name: str,
    row_number: int,
    text: str,
    rule_type: str,
    product_name: str | None = None,
    numeric_value: float | None = None,
    unit: str | None = None
) -> None:
    normalized_name = normalize_product_name(
        product_name or sheet_name,
        sheet_name
    )

    connection.execute(
        """
        INSERT INTO price_rules (
            product_name,
            normalized_name,
            rule_type,
            rule_text,
            numeric_value,
            unit,
            sheet_name,
            row_number,
            original_text
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            product_name,
            normalized_name,
            rule_type,
            text,
            numeric_value,
            unit,
            sheet_name,
            row_number,
            text,
        )
    )


# =========================================================
# 시트별 파서
# =========================================================

def parse_banner_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0
    current_section: str | None = None
    current_material: str | None = None

    for row in range(1, ws.max_row + 1):
        name = normalize_text(ws.cell(row, 2).value)
        specification = normalize_text(
            ws.cell(row, 3).value
        )
        price = to_int_price(
            ws.cell(row, 4).value
        )

        if name in {
            "배너",
            "미니배너",
            "배너대",
        }:
            current_section = name
            current_material = None
            continue

        if name.startswith(NOTE_PREFIXES):
            insert_note(
                connection,
                ws.title,
                row,
                name,
                product_name=current_section
            )
            continue

        if name:
            current_material = name

        if price is None:
            continue

        product_name = (
            current_section
            or ws.title
        )

        width_mm, height_mm = extract_size(
            specification,
            default_unit="mm"
        )

        item = PriceItem(
            product_name=product_name,
            normalized_name=normalize_product_name(
                product_name,
                ws.title
            ),
            category=current_section,
            specification=specification or None,
            width_mm=width_mm,
            height_mm=height_mm,
            material=current_material,
            unit="개",
            unit_price=price,
            total_price=price,
            vat_included=1,
            sheet_name=ws.title,
            row_number=row,
            column_number=4,
            original_text=row_original_text(ws, row),
        )

        insert_price_item(connection, item)
        inserted += 1

    return inserted


def parse_shoulder_strap_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    specification = normalize_text(
        ws.cell(4, 2).value
    )
    material = normalize_text(
        ws.cell(4, 3).value
    )

    width_mm, height_mm = extract_size(
        specification
    )

    for row in range(4, ws.max_row + 1):
        price = to_int_price(
            ws.cell(row, 4).value
        )

        note = normalize_text(
            ws.cell(row, 5).value
        )

        if price is None:
            continue

        quantity_min = None
        quantity_max = None

        if "10개이상" in compact_text(note):
            quantity_min = 10

        elif "10개미만" in compact_text(note):
            quantity_min = 5
            quantity_max = 9

        elif "5개미만" in compact_text(note):
            quantity_min = 1
            quantity_max = 4

        item = PriceItem(
            product_name="어깨띠",
            normalized_name="어깨띠",
            category="어깨띠",
            specification=specification,
            width_mm=width_mm,
            height_mm=height_mm,
            material=material,
            quantity_min=quantity_min,
            quantity_max=quantity_max,
            unit="개",
            unit_price=price,
            total_price=price,
            vat_included=1,
            sheet_name=ws.title,
            row_number=row,
            column_number=4,
            original_text=row_original_text(ws, row),
        )

        insert_price_item(connection, item)
        inserted += 1

        if note:
            insert_rule(
                connection,
                ws.title,
                row,
                note,
                "quantity_condition",
                "어깨띠"
            )

    return inserted


def parse_banner_matrix_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    """
    현수막 시트:
    - B3:P10의 가로×세로 가격 매트릭스
    - B14:M16의 시청 현수막
    """
    inserted = 0

    # 가로 헤더: E4:P4
    width_headers: dict[int, float] = {}

    for col in range(5, 17):
        width_cm = to_float(
            ws.cell(4, col).value
        )

        if width_cm is not None:
            width_headers[col] = width_cm

    # 세로 범위: C5:C10
    for row in range(5, 11):
        height_min_cm, height_max_cm = parse_range(
            ws.cell(row, 3).value
        )

        if height_min_cm is None:
            continue

        for col, width_cm in width_headers.items():
            price = to_int_price(
                ws.cell(row, col).value
            )

            if price is None:
                continue

            item = PriceItem(
                product_name="현수막",
                normalized_name="현수막",
                category="현수막",
                specification=(
                    f"가로 {width_cm:g}cm, "
                    f"세로 {height_min_cm:g}"
                    + (
                        f"~{height_max_cm:g}cm"
                        if height_max_cm != height_min_cm
                        else "cm"
                    )
                ),
                width_mm=width_cm * 10,
                width_mm_min=width_cm * 10,
                width_mm_max=width_cm * 10,
                height_mm=(
                    height_min_cm * 10
                    if height_min_cm == height_max_cm
                    else None
                ),
                height_mm_min=height_min_cm * 10,
                height_mm_max=height_max_cm * 10,
                unit="장",
                unit_price=price,
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=col,
                original_text=row_original_text(
                    ws,
                    row,
                    2,
                    16
                ),
            )

            insert_price_item(connection, item)
            inserted += 1

    # 시청 현수막
    for row in range(15, 17):
        location = normalize_text(
            ws.cell(row, 2).value
        )

        size_text = normalize_text(
            ws.cell(row, 5).value
        )

        finishing = normalize_text(
            ws.cell(row, 8).value
        )

        included_price = to_int_price(
            ws.cell(row, 11).value
        )

        excluded_price = to_int_price(
            ws.cell(row, 13).value
        )

        width_mm, height_mm = extract_size(
            size_text,
            default_unit="m"
        )

        for label, price, installation in (
            ("철거 포함", included_price, "철거 포함"),
            ("철거 제외", excluded_price, "철거 제외"),
        ):
            if price is None:
                continue

            item = PriceItem(
                product_name="시청 현수막",
                normalized_name="현수막",
                category="시청 현수막",
                specification=" / ".join(
                    value
                    for value in (
                        location,
                        size_text,
                        finishing,
                        installation,
                    )
                    if value
                ),
                width_mm=width_mm,
                height_mm=height_mm,
                material="현수막",
                unit="장",
                unit_price=price,
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=(
                    11 if label == "철거 포함" else 13
                ),
                original_text=row_original_text(ws, row),
            )

            insert_price_item(connection, item)
            inserted += 1

    # 참고사항/규칙
    for row in range(18, ws.max_row + 1):
        text = normalize_text(
            ws.cell(row, 2).value
        )

        if not text:
            continue

        rule_type = "note"

        if "철거비" in text:
            rule_type = "removal_fee"
        elif "디자인비" in text:
            rule_type = "design_fee"
        elif "게릴라" in text:
            rule_type = "supplier_unit_price"
        elif "게시대" in text:
            rule_type = "supplier_unit_price"

        insert_rule(
            connection,
            ws.title,
            row,
            text,
            rule_type,
            "현수막"
        )

    return inserted


def parse_business_card_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    # 일반코팅: 수량 B, 면 C, 판매단가 F
    # 고급지: 수량 G, 면 H, 판매단가 K
    blocks = [
        {
            "product": "일반코팅 명함",
            "material": "일반코팅",
            "quantity_col": 2,
            "side_col": 3,
            "price_col": 6,
            "start_row": 4,
            "end_row": 17,
        },
        {
            "product": "고급지 명함",
            "material": "고급지",
            "quantity_col": 7,
            "side_col": 8,
            "price_col": 11,
            "start_row": 4,
            "end_row": 17,
        },
        {
            "product": "명함",
            "material": None,
            "quantity_col": 19,
            "side_col": 20,
            "price_col": 22,
            "start_row": 4,
            "end_row": 21,
        },
    ]

    for block in blocks:
        last_quantity: float | None = None

        for row in range(
            block["start_row"],
            block["end_row"] + 1
        ):
            quantity, _ = extract_quantity(
                ws.cell(
                    row,
                    block["quantity_col"]
                ).value
            )

            if quantity is not None:
                last_quantity = quantity

            side = normalize_text(
                ws.cell(
                    row,
                    block["side_col"]
                ).value
            )

            price = to_int_price(
                ws.cell(
                    row,
                    block["price_col"]
                ).value
            )

            if (
                price is None
                or last_quantity is None
                or side not in {"단면", "양면"}
            ):
                continue

            item = PriceItem(
                product_name=block["product"],
                normalized_name="명함",
                category="명함",
                specification=" / ".join(
                    value
                    for value in (
                        block["material"],
                        side,
                        f"{last_quantity:g}매",
                    )
                    if value
                ),
                material=block["material"],
                print_side=side,
                quantity=last_quantity,
                quantity_min=last_quantity,
                quantity_max=last_quantity,
                unit="매",
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=block["price_col"],
                original_text=row_original_text(ws, row),
            )

            insert_price_item(connection, item)
            inserted += 1

    # 점자명함: 수량 L, 형압 O, 엠보 Q
    for row in range(4, 9):
        quantity, unit = extract_quantity(
            ws.cell(row, 12).value
        )

        if quantity is None:
            continue

        for material, col in (
            ("점자 형압", 15),
            ("점자 엠보", 17),
        ):
            price = to_int_price(
                ws.cell(row, col).value
            )

            if price is None:
                continue

            item = PriceItem(
                product_name="점자명함",
                normalized_name="명함",
                category="점자명함",
                specification=(
                    f"{material} / {quantity:g}{unit or '매'}"
                ),
                material=material,
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit=unit or "매",
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=col,
                original_text=row_original_text(ws, row),
            )

            insert_price_item(connection, item)
            inserted += 1

    # 카드명함 25행부터: 판매가 컬럼들
    card_material_columns = {
        5: "화이트카드",
        7: "골드",
        9: "실버",
        11: "골드펄",
        13: "실버펄",
        15: "실버펄+은테",
    }

    last_quantity = None

    for row in range(24, min(ws.max_row, 50) + 1):
        quantity, unit = extract_quantity(
            ws.cell(row, 2).value
        )

        if quantity is not None:
            last_quantity = quantity

        side = normalize_text(
            ws.cell(row, 3).value
        )

        if side not in {"단면", "양면"}:
            continue

        for price_col, material in card_material_columns.items():
            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if price is None or last_quantity is None:
                continue

            item = PriceItem(
                product_name="카드명함",
                normalized_name="명함",
                category="카드명함",
                specification=(
                    f"{material} / {side} / "
                    f"{last_quantity:g}{unit or '매'}"
                ),
                material=material,
                print_side=side,
                quantity=last_quantity,
                quantity_min=last_quantity,
                quantity_max=last_quantity,
                unit=unit or "매",
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(ws, row),
            )

            insert_price_item(connection, item)
            inserted += 1

    # 후가공/설명
    for row in range(9, 23):
        text = normalize_text(
            ws.cell(row, 12).value
        )

        if text:
            insert_note(
                connection,
                ws.title,
                row,
                text,
                note_type="finishing",
                product_name="명함"
            )

    return inserted

def parse_flyer_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    blocks = [
        # 품목, 용지열, 수량열, 판매가열, 인쇄면, 시작행, 끝행
        ("A4 전단지", 2, 3, 6, "단면", 4, 11),
        ("A4 전단지", 2, 3, 9, "양면", 4, 11),

        ("A4 전단지", 13, 14, 16, "단면", 5, 11),
        ("A4 전단지", 18, 19, 21, "양면", 5, 11),

        ("A4 전단지", 25, 26, 28, "단면", 6, 11),
        ("A4 전단지", 25, 26, 31, "양면", 6, 11),

        ("A4 전단지", 35, 36, 38, "단면", 6, 11),
        ("A4 전단지", 35, 36, 41, "양면", 6, 11),
    ]

    for (
        product_name,
        paper_col,
        quantity_col,
        price_col,
        print_side,
        start_row,
        end_row,
    ) in blocks:
        last_paper = ""

        for row in range(
            start_row,
            min(end_row, ws.max_row) + 1
        ):
            paper_text = normalize_text(
                ws.cell(row, paper_col).value
            )

            if paper_text:
                last_paper = paper_text

            quantity, unit = extract_quantity(
                ws.cell(row, quantity_col).value
            )

            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if (
                quantity is None
                or quantity <= 0
                or price is None
            ):
                continue

            item = PriceItem(
                product_name=product_name,
                normalized_name="전단지",
                category="전단지",
                specification=(
                    f"A4 / {last_paper or '용지 미확인'} / "
                    f"{print_side} / {quantity:g}매"
                ),
                paper=last_paper or None,
                print_side=print_side,
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit=unit or "매",
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(
                    ws,
                    row
                ),
                confidence=1.0,
                review_required=0,
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    return inserted

def parse_poster_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    # A2 소량/기본 단가: 수량 B, 단가 E
    last_quantity: float | None = None

    for row in range(5, 24):
        quantity, _ = extract_quantity(
            ws.cell(row, 2).value
        )

        if quantity is not None:
            last_quantity = quantity

        price = to_int_price(
            ws.cell(row, 5).value
        )

        if price is None or last_quantity is None:
            continue

        item = PriceItem(
            product_name="A2 포스터",
            normalized_name="포스터",
            category="포스터",
            specification=(
                f"A2 / 아트지 150g / 단면 / "
                f"{last_quantity:g}매"
            ),
            paper="아트지 150g",
            print_side="단면",
            quantity=last_quantity,
            quantity_min=last_quantity,
            quantity_max=last_quantity,
            unit="매",
            total_price=price,
            vat_included=1,
            sheet_name=ws.title,
            row_number=row,
            column_number=5,
            original_text=row_original_text(ws, row),
        )

        insert_price_item(connection, item)
        inserted += 1

    # 우측 애즈랜드 A2/A1 단가
    blocks = [
        (9, 11, "A2", "아트지 150g"),
        (9, 14, "A1", "아트지 150g"),
        (16, 18, "A2", "아트지 150g"),
        (16, 21, "A1", "아트지 150g"),
    ]

    for quantity_col, price_col, size, paper in blocks:
        for row in range(5, ws.max_row + 1):
            quantity, _ = extract_quantity(
                ws.cell(row, quantity_col).value
            )

            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if quantity is None or price is None:
                continue

            item = PriceItem(
                product_name=f"{size} 포스터",
                normalized_name="포스터",
                category="포스터",
                specification=(
                    f"{size} / {paper} / 단면 / "
                    f"{quantity:g}매"
                ),
                paper=paper,
                print_side="단면",
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit="매",
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(ws, row),
            )

            insert_price_item(connection, item)
            inserted += 1

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            text = normalize_text(
                ws.cell(row, col).value
            )

            if (
                text.startswith(NOTE_PREFIXES)
                or "경우" in text
                or "가능" in text
            ):
                insert_note(
                    connection,
                    ws.title,
                    row,
                    text,
                    product_name="포스터"
                )

    return inserted


def parse_board_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    """
    포맥스·아크릴 시트의 실제 원자재 판매가를 추출.
    업체·원가 설명은 price_notes로 이동.
    """
    inserted = 0

    sections = [
        # material, start_row, end_row, thickness_col, size_price pairs
        ("포맥스", 5, 10, 9, [(10, "1220*2440"), (11, "900*1800")]),
        ("포맥스", 5, 10, 12, [(13, "1220*2440"), (14, "900*1800")]),
        ("검정색 포맥스", 13, 14, 12, [(13, "1220*2440"), (14, "900*1800")]),
        ("아크릴", 17, 20, 12, [(13, "1220*2440"), (14, "900*1800")]),
    ]

    for material, start_row, end_row, thickness_col, price_columns in sections:
        for row in range(start_row, end_row + 1):
            thickness_text = normalize_text(
                ws.cell(row, thickness_col).value
            )

            thickness = extract_thickness(
                thickness_text
            )

            if thickness is None:
                continue

            for price_col, size_text in price_columns:
                price = to_int_price(
                    ws.cell(row, price_col).value
                )

                if price is None:
                    continue

                width_mm, height_mm = extract_size(
                    size_text
                )

                normalized = (
                    "아크릴"
                    if "아크릴" in material
                    else "포맥스"
                )

                item = PriceItem(
                    product_name=material,
                    normalized_name=normalized,
                    category=normalized,
                    specification=(
                        f"{material} / {thickness:g}T / "
                        f"{size_text}"
                    ),
                    width_mm=width_mm,
                    height_mm=height_mm,
                    thickness_mm=thickness,
                    material=material,
                    unit="장",
                    unit_price=price,
                    total_price=price,
                    vat_included=None,
                    sheet_name=ws.title,
                    row_number=row,
                    column_number=price_col,
                    original_text=row_original_text(ws, row),
                )

                insert_price_item(connection, item)
                inserted += 1

    for row in range(1, ws.max_row + 1):
        text = row_original_text(ws, row)

        if (
            "UV" in text
            or "인쇄" in text
            or "재단" in text
            or "프라이머" in text
        ):
            insert_note(
                connection,
                ws.title,
                row,
                text,
                product_name="포맥스/아크릴"
            )

    return inserted


def find_price_columns(
    ws: Worksheet
) -> list[tuple[int, int]]:
    """
    (header_row, price_column)
    """
    result: list[tuple[int, int]] = []

    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, ws.max_column + 1):
            compact = compact_text(
                ws.cell(row, col).value
            )

            if compact in {
                compact_text(value)
                for value in SUPPORTED_PRICE_HEADERS
            }:
                result.append((row, col))

    return result


def infer_product_context(
    ws: Worksheet,
    header_row: int,
    price_col: int
) -> str:
    candidates: list[str] = []

    # 같은 열 위쪽
    for row in range(header_row - 1, 0, -1):
        text = normalize_text(
            ws.cell(row, price_col).value
        )

        if text:
            candidates.append(text)
            break

    # 헤더 왼쪽 위
    for col in range(price_col - 1, 0, -1):
        text = normalize_text(
            ws.cell(header_row, col).value
        )

        if text:
            candidates.append(text)
            break

    # 시트명
    candidates.append(ws.title)

    for candidate in candidates:
        if not is_blocked_product_text(candidate):
            return candidate

    return ws.title


def generic_table_parser(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    """
    전단지·리플릿·카다로그·봉투 등 복합 시트용 보수적 파서.

    단가 헤더 아래의 숫자만 후보로 삼고,
    같은 행의 수량·규격·인쇄면 조건을 찾는다.
    구매 조건이 없으면 review_items로 보낸다.
    """
    inserted = 0
    price_headers = find_price_columns(ws)

    for header_row, price_col in price_headers:
        # 범용 파서에서는 주변 셀의 "단면", "판매가",
        # "인쇄비2배", 업체명 등을 품목명으로 오인할 수 있다.
        # 따라서 기본 품목명은 시트명으로 고정한다.
        product_context = ws.title.strip()

        normalized_name = normalize_product_name(
            product_context,
            ws.title
        )

        empty_count = 0

        for row in range(
            header_row + 1,
            ws.max_row + 1
        ):
            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if price is None:
                empty_count += 1

                if empty_count >= 8:
                    break

                continue

            empty_count = 0

            row_values = [
                normalize_text(
                    ws.cell(row, col).value
                )
                for col in range(
                    max(1, price_col - 6),
                    min(ws.max_column, price_col + 2) + 1
                )
            ]

            joined = " / ".join(
                value
                for value in row_values
                if value
            )

            # 수량/규격/인쇄면 추출
            quantity = None
            unit = None

            for col in range(
                    max(1, price_col - 6),
                    price_col
            ):
                quantity_candidate, unit_candidate = (
                    extract_quantity_strict(
                        ws.cell(row, col).value
                    )
                )

                if quantity_candidate is not None:
                    quantity = quantity_candidate
                    unit = unit_candidate
                    break

            width_mm, height_mm = extract_size(joined)
            thickness = extract_thickness(joined)
            paper = extract_paper(joined)
            print_side = infer_print_side(joined)

            material = None
            for keyword in (
                "아트지",
                "스노우지",
                "모조지",
                "랑데뷰",
                "휘라레",
                "부직포",
                "패트지",
                "메쉬",
                "포맥스",
                "아크릴",
            ):
                if keyword in joined:
                    material = keyword
                    break

            specification_parts = [
                value
                for value in (
                    joined or None,
                    (
                        f"{quantity:g}{unit or ''}"
                        if quantity is not None
                        else None
                    ),
                )
                if value
            ]

            item = PriceItem(
                product_name=product_context,
                normalized_name=normalized_name,
                category=ws.title.strip(),
                specification=(
                    " / ".join(specification_parts)
                    if specification_parts
                    else None
                ),
                width_mm=width_mm,
                height_mm=height_mm,
                thickness_mm=thickness,
                material=material,
                paper=paper,
                print_side=print_side,
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit=unit,
                total_price=price,
                vat_included=(
                    1
                    if "부가세포함" in compact_text(
                        row_original_text(
                            ws,
                            max(1, header_row - 3),
                            1,
                            ws.max_column
                        )
                    )
                    else None
                ),
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(ws, row),
                confidence=0.65,
                review_required=1,
            )

            before = connection.total_changes
            insert_price_item(connection, item)

            if connection.total_changes > before:
                inserted += 1

    # 주석/참고문 저장
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            text = normalize_text(
                ws.cell(row, col).value
            )

            if (
                text.startswith(NOTE_PREFIXES)
                or "참고" in text
                or "주의" in text
                or "별도" in text
                or "포함" in text
            ):
                insert_note(
                    connection,
                    ws.title,
                    row,
                    text
                )

    return inserted



# =========================================================
# 전체 시트용 명시적 판매가 파서
# =========================================================

EXPLICIT_SALE_HEADERS = {
    "단가",
    "판매가",
    "판매단가",
    "공급가",
    "공급금액",
    "금액",
    "디자인비+인쇄비",
    "인쇄비2배",
    "페이지당단가",
}

NON_SALE_HEADERS = {
    "원가",
    "마진",
    "하청가",
    "이전단가",
    "차이",
    "애즈",
    "와우프레스",
}


def get_merged_top_left_value(
    ws: Worksheet,
    row: int,
    col: int
) -> Any:
    """
    병합 셀 내부 좌표가 들어오면 병합 영역의 좌상단 값을 반환한다.
    """
    coordinate = ws.cell(row, col).coordinate

    for merged_range in ws.merged_cells.ranges:
        if coordinate in merged_range:
            return ws.cell(
                merged_range.min_row,
                merged_range.min_col
            ).value

    return ws.cell(row, col).value


def find_header_above(
    ws: Worksheet,
    row: int,
    col: int,
    max_distance: int = 5
) -> tuple[int | None, str | None]:
    """
    현재 숫자 셀 위쪽에서 가격 의미 헤더를 찾는다.
    """
    for target_row in range(
        row - 1,
        max(0, row - max_distance - 1),
        -1
    ):
        value = normalize_text(
            get_merged_top_left_value(
                ws,
                target_row,
                col
            )
        )

        compact = compact_text(value)

        if not compact:
            continue

        if compact in {
            compact_text(header)
            for header in EXPLICIT_SALE_HEADERS
        }:
            return target_row, value

        if compact in {
            compact_text(header)
            for header in NON_SALE_HEADERS
        }:
            return target_row, value

    return None, None


def find_context_in_row(
    ws: Worksheet,
    row: int,
    price_col: int
) -> tuple[
    float | None,
    str | None,
    str | None,
    str | None
]:
    """
    판매가 셀 왼쪽에서 수량, 단위, 용지/재질, 인쇄면을 찾는다.
    """
    quantity: float | None = None
    unit: str | None = None
    material: str | None = None
    print_side: str | None = None

    values: list[str] = []

    for col in range(
        max(1, price_col - 8),
        price_col
    ):
        raw = get_merged_top_left_value(
            ws,
            row,
            col
        )

        text_value = normalize_text(raw)

        if text_value:
            values.append(text_value)

        if quantity is None:
            q, u = extract_quantity(raw)

            # 범용 파서이지만, 가격 헤더가 명확한 열에서만 사용한다.
            # 숫자 단독 수량도 허용하되 20g~500g 같은 평량 문구는 제외한다.
            if q is not None:
                compact = compact_text(text_value)

                if not (
                    compact.endswith("g")
                    or "페이지" in compact
                    or compact.endswith("p")
                ):
                    quantity = q
                    unit = u

    joined = " / ".join(values)

    for keyword in (
        "아트지",
        "스노우지",
        "모조지",
        "백상지",
        "랑데뷰",
        "휘라레",
        "반누보",
        "레자크",
        "크라프트",
        "부직포",
        "패트지",
        "메쉬",
        "포맥스",
        "아크릴",
        "골판지",
        "허니콤",
    ):
        if keyword in joined:
            material = keyword
            break

    print_side = infer_print_side(joined)

    return (
        quantity,
        unit,
        material,
        print_side
    )


def find_product_and_specification(
    ws: Worksheet,
    row: int,
    price_col: int
) -> tuple[str, str]:
    """
    시트명과 가격 주변 셀을 조합해 사람이 읽을 수 있는 품목/규격을 만든다.
    """
    product_name = ws.title.strip()

    context_parts: list[str] = []

    # 같은 행의 왼쪽 정보
    for col in range(
        max(1, price_col - 8),
        price_col
    ):
        text_value = normalize_text(
            get_merged_top_left_value(
                ws,
                row,
                col
            )
        )

        if not text_value:
            continue

        compact = compact_text(text_value)

        if compact in {
            compact_text(header)
            for header in (
                EXPLICIT_SALE_HEADERS
                | NON_SALE_HEADERS
            )
        }:
            continue

        if to_int_price(text_value) is not None:
            continue

        context_parts.append(text_value)

    # 위쪽의 블록 제목/용지/규격 정보
    for target_row in range(
        row - 1,
        max(0, row - 6),
        -1
    ):
        for target_col in range(
            max(1, price_col - 3),
            min(ws.max_column, price_col + 1) + 1
        ):
            text_value = normalize_text(
                get_merged_top_left_value(
                    ws,
                    target_row,
                    target_col
                )
            )

            if not text_value:
                continue

            compact = compact_text(text_value)

            if compact in {
                compact_text(header)
                for header in (
                    EXPLICIT_SALE_HEADERS
                    | NON_SALE_HEADERS
                )
            }:
                continue

            if to_int_price(text_value) is not None:
                continue

            if len(text_value) <= 80:
                context_parts.append(text_value)

    # 중복 제거
    unique_parts = list(
        dict.fromkeys(
            part
            for part in context_parts
            if part
        )
    )

    specification = " / ".join(
        unique_parts[:12]
    )

    return product_name, specification


def parse_explicit_sale_columns(
    ws: Worksheet,
    connection: sqlite3.Connection,
    *,
    force_product_name: str | None = None,
    vat_included: int | None = None
) -> int:
    """
    복잡한 표에서 '단가/판매가/금액/공급금액/디자인비+인쇄비'처럼
    판매 의미가 명시된 열만 추출한다.

    원가, 마진, 하청가, 이전 단가는 제외한다.
    """
    inserted = 0

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            price = to_int_price(
                ws.cell(row, col).value
            )

            if price is None:
                continue

            header_row, header_text = find_header_above(
                ws,
                row,
                col,
                max_distance=5
            )

            if header_text is None:
                continue

            header_compact = compact_text(
                header_text
            )

            if header_compact in {
                compact_text(header)
                for header in NON_SALE_HEADERS
            }:
                continue

            if header_compact not in {
                compact_text(header)
                for header in EXPLICIT_SALE_HEADERS
            }:
                continue

            (
                quantity,
                unit,
                material,
                print_side
            ) = find_context_in_row(
                ws,
                row,
                col
            )

            product_name, specification = (
                find_product_and_specification(
                    ws,
                    row,
                    col
                )
            )

            if force_product_name:
                product_name = force_product_name

            width_mm, height_mm = extract_size(
                specification
            )

            thickness = extract_thickness(
                specification
            )

            paper = extract_paper(
                specification
            )

            normalized_name = normalize_product_name(
                product_name,
                ws.title
            )

            # '단가'는 개당 단가, 나머지는 대체로 해당 수량의 전체 금액이다.
            if header_compact in {
                compact_text("단가"),
                compact_text("판매단가"),
                compact_text("페이지당단가"),
            }:
                unit_price = price

                total_price = (
                    int(round(price * quantity))
                    if quantity is not None
                    else None
                )

            else:
                unit_price = None
                total_price = price

            item = PriceItem(
                product_name=product_name,
                normalized_name=normalized_name,
                category=ws.title.strip(),
                specification=(
                    specification
                    or f"{header_text} 기준"
                ),
                width_mm=width_mm,
                height_mm=height_mm,
                thickness_mm=thickness,
                material=material,
                paper=paper,
                print_side=print_side,
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit=unit,
                unit_price=unit_price,
                total_price=total_price,
                vat_included=vat_included,
                sheet_name=ws.title,
                row_number=row,
                column_number=col,
                original_text=row_original_text(
                    ws,
                    row
                ),
                confidence=0.9,
                review_required=0,
            )

            before = connection.total_changes

            insert_price_item(
                connection,
                item
            )

            if connection.total_changes > before:
                inserted += 1

    return inserted


def parse_overpass_banner_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    last_bridge = ""
    last_address = ""

    for row in range(4, ws.max_row + 1):
        bridge = normalize_text(
            ws.cell(row, 3).value
        )

        address = normalize_text(
            ws.cell(row, 4).value
        )

        if bridge:
            last_bridge = bridge

        if address:
            last_address = address

        direction = normalize_text(
            ws.cell(row, 5).value
        )

        route = normalize_text(
            ws.cell(row, 6).value
        )

        size_text = normalize_text(
            ws.cell(row, 7).value
        )

        unit_price = to_int_price(
            ws.cell(row, 10).value
        )

        supply_price = to_int_price(
            ws.cell(row, 12).value
        )

        vat_price = to_int_price(
            ws.cell(row, 13).value
        )

        if not size_text:
            continue

        width_mm, height_mm = extract_size(
            size_text
        )

        price = (
            vat_price
            or supply_price
            or unit_price
        )

        if price is None:
            continue

        item = PriceItem(
            product_name="육교현수막",
            normalized_name="현수막",
            category="육교현수막",
            specification=" / ".join(
                value
                for value in (
                    last_bridge,
                    last_address,
                    direction,
                    route,
                    size_text,
                )
                if value
            ),
            width_mm=width_mm,
            height_mm=height_mm,
            unit="장",
            unit_price=price,
            total_price=price,
            vat_included=(
                1 if vat_price is not None else 0
            ),
            sheet_name=ws.title,
            row_number=row,
            column_number=(
                13
                if vat_price is not None
                else 12
            ),
            original_text=row_original_text(
                ws,
                row
            ),
        )

        insert_price_item(
            connection,
            item
        )

        inserted += 1

    return inserted


def parse_id_card_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    for row in range(2, 5):
        product = normalize_text(
            ws.cell(row, 2).value
        )

        price = to_int_price(
            ws.cell(row, 3).value
        )

        note = normalize_text(
            ws.cell(row, 4).value
        )

        if product and price is not None:
            item = PriceItem(
                product_name=product,
                normalized_name="사원증",
                category="사원증",
                specification=note or product,
                unit="개",
                unit_price=price,
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=3,
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    for row in (2, 3):
        condition = normalize_text(
            ws.cell(row, 6).value
        )

        price = to_int_price(
            ws.cell(row, 7).value
        )

        if not condition or price is None:
            continue

        quantity_min = None
        quantity_max = None

        compact = compact_text(condition)

        if "10개미만" in compact:
            quantity_min = 1
            quantity_max = 9

        elif "10개이상" in compact:
            quantity_min = 10

        item = PriceItem(
            product_name="사원증 세트",
            normalized_name="사원증",
            category="사원증세트",
            specification=condition,
            quantity_min=quantity_min,
            quantity_max=quantity_max,
            unit="세트",
            unit_price=price,
            total_price=price,
            vat_included=1,
            sheet_name=ws.title,
            row_number=row,
            column_number=7,
            original_text=row_original_text(
                ws,
                row
            ),
        )

        insert_price_item(
            connection,
            item
        )

        inserted += 1

    return inserted


def parse_infographic_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0
    current_product = ""

    for row in range(1, ws.max_row + 1):
        label = normalize_text(
            ws.cell(row, 2).value
        )

        if label in {
            "인포그래픽",
            "PPT",
        }:
            current_product = label
            continue

        price = to_int_price(
            ws.cell(row, 3).value
        )

        if price is None:
            continue

        item = PriceItem(
            product_name=(
                current_product
                or "인포그래픽"
            ),
            normalized_name="인포그래픽",
            category=current_product or "인포그래픽",
            specification=(
                label
                or "난이도별 페이지당"
            ),
            unit="페이지",
            unit_price=price,
            total_price=price,
            vat_included=1,
            sheet_name=ws.title,
            row_number=row,
            column_number=3,
            original_text=row_original_text(
                ws,
                row
            ),
        )

        insert_price_item(
            connection,
            item
        )

        inserted += 1

    return inserted


def parse_binding_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0
    current_part = ""

    for row in range(2, 6):
        part = normalize_text(
            ws.cell(row, 2).value
        )

        if part:
            current_part = part

        for supplier, color_col, price_col in (
            ("열린문디자인", 3, 4),
            ("준디자인", 5, 6),
        ):
            color = normalize_text(
                ws.cell(row, color_col).value
            )

            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if not color or price is None:
                continue

            item = PriceItem(
                product_name="책제본",
                normalized_name="책제본",
                category=current_part or "책제본",
                specification=(
                    f"{current_part} / "
                    f"{color} / {supplier}"
                ),
                color=color,
                unit=(
                    "페이지"
                    if current_part == "내지"
                    else "부"
                ),
                unit_price=price,
                total_price=price,
                vat_included=None,
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    for row in range(13, ws.max_row + 1):
        text_value = row_original_text(
            ws,
            row
        )

        if text_value:
            insert_note(
                connection,
                ws.title,
                row,
                text_value,
                product_name="책제본"
            )

    return inserted


def parse_movie_supplier_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    quantity_headers: dict[int, float] = {}

    for col in range(6, 10):
        quantity, _ = extract_quantity(
            ws.cell(2, col).value
        )

        if quantity is not None:
            quantity_headers[col] = quantity

    current_size = ""

    for row in range(3, 7):
        size = normalize_text(
            ws.cell(row, 3).value
        )

        if size:
            current_size = size

        tape = normalize_text(
            ws.cell(row, 4).value
        )

        for col, quantity in quantity_headers.items():
            price = to_int_price(
                ws.cell(row, col).value
            )

            if price is None:
                continue

            item = PriceItem(
                product_name="옵셋봉투",
                normalized_name="봉투",
                category="영화인재",
                specification=" / ".join(
                    value
                    for value in (
                        current_size,
                        tape,
                    )
                    if value
                ),
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit="매",
                total_price=price,
                vat_included=0,
                sheet_name=ws.title,
                row_number=row,
                column_number=col,
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    for row, product_name, unit in (
        (8, "초대장봉투", "개"),
        (9, "초대장 스티커", "개"),
        (12, "이중봉투", "개"),
        (20, "상장지", "장"),
    ):
        price = to_int_price(
            ws.cell(row, 4).value
        )

        if price is None:
            continue

        item = PriceItem(
            product_name=product_name,
            normalized_name=normalize_product_name(
                product_name,
                ws.title
            ),
            category=ws.title.strip(),
            specification=normalize_text(
                ws.cell(row, 3).value
            ) or product_name,
            unit=unit,
            unit_price=price,
            total_price=price,
            vat_included=0,
            sheet_name=ws.title,
            row_number=row,
            column_number=4,
            original_text=row_original_text(
                ws,
                row
            ),
        )

        insert_price_item(
            connection,
            item
        )

        inserted += 1

    return inserted


def parse_certificate_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    blocks = [
        {
            "product": "상장지 매직콤마120g 은박",
            "quantity_col": 2,
            "unit_price_col": 4,
            "total_price_col": 5,
            "start_row": 5,
            "end_row": 15,
        },
        {
            "product": "상장지 매직코튼120g 금박",
            "quantity_col": 8,
            "unit_price_col": 10,
            "total_price_col": 11,
            "start_row": 5,
            "end_row": 15,
        },
    ]

    for block in blocks:
        for row in range(
            block["start_row"],
            block["end_row"] + 1
        ):
            quantity = to_float(
                ws.cell(
                    row,
                    block["quantity_col"]
                ).value
            )

            unit_price = to_int_price(
                ws.cell(
                    row,
                    block["unit_price_col"]
                ).value
            )

            total_price = to_int_price(
                ws.cell(
                    row,
                    block["total_price_col"]
                ).value
            )

            if (
                quantity is None
                or (
                    unit_price is None
                    and total_price is None
                )
            ):
                continue

            item = PriceItem(
                product_name=block["product"],
                normalized_name="상장지",
                category="상장지",
                specification=(
                    f"{block['product']} / "
                    f"{quantity:g}장"
                ),
                paper=(
                    "매직콤마 120g"
                    if "콤마" in block["product"]
                    else "매직코튼 120g"
                ),
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit="장",
                unit_price=unit_price,
                total_price=(
                    total_price
                    or (
                        int(round(unit_price * quantity))
                        if unit_price is not None
                        else None
                    )
                ),
                vat_included=0,
                sheet_name=ws.title,
                row_number=row,
                column_number=block[
                    "total_price_col"
                ],
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    return inserted


def parse_form_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    blocks = [
        ("양식지 단면", 1, 2, 3, 5, 14, "매"),
        ("양식지 양면", 5, 6, 7, 5, 14, "매"),
        ("NCR지", 10, None, 11, 5, 9, "권"),
        ("인감 양식지", 19, 20, 21, 3, 7, "매"),
        ("양식지 단면", 24, 25, 26, 3, 7, "매"),
        ("양식지 양면", 29, 30, 31, 3, 7, "매"),
        ("양식지 단면", 19, 20, 21, 13, 22, "매"),
        ("양식지 양면", 23, 24, 25, 13, 22, "매"),
        ("양식지 단면", 27, 28, 29, 13, 17, "매"),
        ("양식지 양면", 32, 33, 34, 13, 17, "매"),
    ]

    for (
        product,
        quantity_col,
        unit_price_col,
        total_price_col,
        start_row,
        end_row,
        unit
    ) in blocks:
        for row in range(
            start_row,
            min(end_row, ws.max_row) + 1
        ):
            quantity = to_float(
                ws.cell(row, quantity_col).value
            )

            unit_price = (
                to_int_price(
                    ws.cell(
                        row,
                        unit_price_col
                    ).value
                )
                if unit_price_col
                else None
            )

            total_price = (
                to_int_price(
                    ws.cell(
                        row,
                        total_price_col
                    ).value
                )
                if total_price_col
                else None
            )

            if (
                quantity is None
                or (
                    unit_price is None
                    and total_price is None
                )
            ):
                continue

            item = PriceItem(
                product_name=product,
                normalized_name="양식지",
                category=product,
                specification=(
                    f"{product} / "
                    f"{quantity:g}{unit}"
                ),
                paper=(
                    "백상지 80g"
                    if "양식지" in product
                    else None
                ),
                print_side=(
                    "양면"
                    if "양면" in product
                    else (
                        "단면"
                        if "단면" in product
                        else None
                    )
                ),
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit=unit,
                unit_price=unit_price,
                total_price=(
                    total_price
                    or (
                        int(
                            round(
                                unit_price
                                * quantity
                            )
                        )
                        if unit_price is not None
                        else None
                    )
                ),
                vat_included=None,
                sheet_name=ws.title,
                row_number=row,
                column_number=(
                    total_price_col
                    or unit_price_col
                ),
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    return inserted


def parse_honeycomb_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    # 2026년 원자재 단가
    for material_col, size_col, price_col in (
        (1, 2, 3),
        (4, 5, 5),
    ):
        current_material = ""

        for row in range(3, 12):
            material = normalize_text(
                ws.cell(row, material_col).value
            )

            if material:
                current_material = material

            size_text = normalize_text(
                ws.cell(row, size_col).value
            )

            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if not current_material or price is None:
                continue

            width_mm, height_mm = extract_size(
                size_text
            )

            thickness = extract_thickness(
                current_material
            )

            item = PriceItem(
                product_name=current_material,
                normalized_name="골지보드",
                category="골지·허니콤보드",
                specification=" / ".join(
                    value
                    for value in (
                        current_material,
                        size_text,
                    )
                    if value
                ),
                width_mm=width_mm,
                height_mm=height_mm,
                thickness_mm=thickness,
                material=current_material,
                unit="장",
                unit_price=price,
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    # 캐드리 수량 구간별 2026년 단가
    quantity_columns = {
        3: 100,
        4: 50,
        5: 1,
    }

    for row in range(14, 30):
        product = normalize_text(
            ws.cell(row, 1).value
        )

        size_text = normalize_text(
            ws.cell(row, 2).value
        )

        if not product:
            continue

        for price_col, quantity in (
            quantity_columns.items()
        ):
            price = to_int_price(
                ws.cell(row, price_col).value
            )

            if price is None:
                continue

            width_mm, height_mm = extract_size(
                size_text
            )

            item = PriceItem(
                product_name=product,
                normalized_name="골지보드",
                category="골지·허니콤보드",
                specification=(
                    f"{product} / {size_text} / "
                    f"{quantity}장 기준"
                ),
                width_mm=width_mm,
                height_mm=height_mm,
                thickness_mm=extract_thickness(
                    product
                ),
                material=product,
                quantity=quantity,
                quantity_min=quantity,
                quantity_max=quantity,
                unit="장",
                unit_price=price,
                total_price=int(
                    round(price * quantity)
                ),
                vat_included=1,
                sheet_name=ws.title,
                row_number=row,
                column_number=price_col,
                original_text=row_original_text(
                    ws,
                    row
                ),
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    # UV/재단 등 명시된 판매가
    inserted += parse_explicit_sale_columns(
        ws,
        connection,
        force_product_name="골지·허니콤보드",
        vat_included=1
    )

    return inserted


def parse_eco_banner_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    inserted = 0

    for col in (
        7,
        12,
        17,
    ):
        text_value = normalize_text(
            ws.cell(15, col).value
        )

        if not text_value:
            continue

        title = text_value.splitlines()[0]

        patterns = [
            (
                "지지대+인쇄내지",
                r"지지대\+인쇄내지\s*=\s*([\d,]+)"
            ),
            (
                "인쇄내지만",
                r"인쇄내지만(?:\s*주문시)?\s*=\s*([\d,]+)"
            ),
            (
                title,
                r"(?:^|\n)\s*([\d,]+)\s*원"
            ),
        ]

        for subtype, pattern in patterns:
            match = re.search(
                pattern,
                text_value
            )

            if not match:
                continue

            price = int(
                match.group(1).replace(",", "")
            )

            item = PriceItem(
                product_name="친환경배너",
                normalized_name="친환경배너",
                category=title,
                specification=(
                    f"{title} / {subtype}"
                ),
                material="친환경 배너",
                unit="세트",
                unit_price=price,
                total_price=price,
                vat_included=1,
                sheet_name=ws.title,
                row_number=15,
                column_number=col,
                original_text=text_value,
            )

            insert_price_item(
                connection,
                item
            )

            inserted += 1

    return inserted

def parse_sheet(
    ws: Worksheet,
    connection: sqlite3.Connection
) -> int:
    """
    현재 열린문디자인 단가표의 20개 시트를 모두 전용 파서로 처리한다.

    등록되지 않은 시트는 조용히 범용 처리하지 않고 오류를 발생시킨다.
    새 시트가 추가되면 반드시 전용 파서를 등록해야 한다.
    """
    title = ws.title.strip()

    parser_map = {
        "현수막": parse_banner_matrix_sheet,
        "육교현수막": parse_overpass_banner_sheet,
        "배너": parse_banner_sheet,
        "어깨띠": parse_shoulder_strap_sheet,
        "사원증": parse_id_card_sheet,
        "인포그래픽": parse_infographic_sheet,
        "책제본": parse_binding_sheet,
        "명함": parse_business_card_sheet,
        "포스터": parse_poster_sheet,
        "상장지": parse_certificate_sheet,
        "양식지": parse_form_sheet,
        "골지,허니콤보드": parse_honeycomb_sheet,
        "포맥스,아크릴": parse_board_sheet,
        "친환경배너": parse_eco_banner_sheet,
        "영화인재": parse_movie_supplier_sheet,
    }

    if title == "제작업체":
        for row in range(
            1,
            ws.max_row + 1
        ):
            text_value = row_original_text(
                ws,
                row
            )

            if text_value:
                insert_note(
                    connection,
                    ws.title,
                    row,
                    text_value,
                    note_type="supplier"
                )

        return 0

    parser = parser_map.get(title)

    if parser is not None:
        return parser(
            ws,
            connection
        )

    if title == "전단지":
        return parse_flyer_sheet(
            ws,
            connection
        )

    # 아래 복합 인쇄 시트는 명시적인 판매가/단가 열만 전부 추출한다.
    if title in {
        "리플릿",
        "카다로그",
        "옵셋봉투",
    }:
        return parse_explicit_sale_columns(
            ws,
            connection,
            force_product_name=title,
            vat_included=(
                1
                if "부가세포함" in compact_text(
                    row_original_text(
                        ws,
                        1
                    )
                )
                else None
            )
        )

    raise RuntimeError(
        "전용 파서가 등록되지 않은 시트입니다: "
        f"{title}"
    )


# =========================================================
# 검증
# =========================================================

def validate_database(
    connection: sqlite3.Connection
) -> None:
    invalid_no_price = connection.execute(
        """
        SELECT COUNT(*)
        FROM price_items
        WHERE
            unit_price IS NULL
            AND total_price IS NULL
        """
    ).fetchone()[0]

    if invalid_no_price:
        raise RuntimeError(
            "가격이 없는 price_items가 존재합니다: "
            f"{invalid_no_price}건"
        )

    blocked_count = 0

    rows = connection.execute(
        """
        SELECT id, product_name
        FROM price_items
        """
    ).fetchall()

    for row in rows:
        if is_blocked_product_text(
            row["product_name"]
        ):
            blocked_count += 1

    if blocked_count:
        raise RuntimeError(
            "헤더·설명으로 판단되는 price_items가 "
            f"{blocked_count}건 존재합니다."
        )


def print_summary(
    connection: sqlite3.Connection
) -> None:
    item_count = connection.execute(
        "SELECT COUNT(*) FROM price_items"
    ).fetchone()[0]

    rule_count = connection.execute(
        "SELECT COUNT(*) FROM price_rules"
    ).fetchone()[0]

    note_count = connection.execute(
        "SELECT COUNT(*) FROM price_notes"
    ).fetchone()[0]

    review_count = connection.execute(
        "SELECT COUNT(*) FROM review_items"
    ).fetchone()[0]

    print()
    print("=" * 80)
    print("price_table.db 생성 완료")
    print("판매 가격 항목:", item_count)
    print("가격 규칙:", rule_count)
    print("참고사항:", note_count)
    print("검토 필요 후보:", review_count)
    print("=" * 80)

    print()
    print("[품목별 저장 건수]")

    rows = connection.execute(
        """
        SELECT
            normalized_name,
            COUNT(*) AS count
        FROM price_items
        GROUP BY normalized_name
        ORDER BY count DESC
        """
    ).fetchall()

    for row in rows:
        print(
            f"  {row['normalized_name']}: "
            f"{row['count']}건"
        )

    print()
    print("[샘플 20건]")

    rows = connection.execute(
        """
        SELECT
            product_name,
            normalized_name,
            specification,
            quantity,
            unit,
            unit_price,
            total_price,
            sheet_name,
            row_number,
            review_required
        FROM price_items
        ORDER BY id
        LIMIT 20
        """
    ).fetchall()

    for row in rows:
        print(dict(row))


# =========================================================
# 실행
# =========================================================

def build_price_database(
    input_path: Path,
    database_path: Path
) -> None:
    if not input_path.exists():
        raise FileNotFoundError(
            f"단가표 Excel이 없습니다: "
            f"{input_path.resolve()}"
        )

    if database_path.exists():
        database_path.unlink()

    workbook = load_workbook(
        input_path,
        data_only=True,
        read_only=False
    )

    connection = connect_database(
        database_path
    )

    create_schema(connection)

    try:
        connection.execute(
            """
            INSERT INTO source_files (
                file_path,
                file_name,
                file_hash,
                indexed_at
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                str(input_path.resolve()),
                input_path.name,
                file_sha256(input_path),
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            )
        )

        for ws in workbook.worksheets:
            print(
                f"[처리] {ws.title}"
            )

            try:
                inserted = parse_sheet(
                    ws,
                    connection
                )

                print(
                    f"       price_items {inserted}건"
                )

            except Exception as error:
                print(
                    f"       실패: "
                    f"{type(error).__name__}: {error}"
                )

                traceback.print_exc()

                connection.execute(
                    """
                    INSERT INTO review_items (
                        sheet_name,
                        reason,
                        original_text
                    )
                    VALUES (?, ?, ?)
                    """,
                    (
                        ws.title,
                        (
                            f"시트 파싱 실패: "
                            f"{type(error).__name__}: {error}"
                        ),
                        None,
                    )
                )

        connection.commit()

        validate_database(
            connection
        )

        print_summary(
            connection
        )

    finally:
        workbook.close()
        connection.close()


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "열린문디자인 단가표 Excel을 "
            "검색용 SQLite DB로 변환합니다."
        )
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help="입력 단가표 Excel 경로"
    )

    parser.add_argument(
        "--db",
        type=Path,
        default=DEFAULT_DB_PATH,
        help="출력 SQLite DB 경로"
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    build_price_database(
        input_path=args.input,
        database_path=args.db
    )


if __name__ == "__main__":
    main()
