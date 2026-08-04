from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================================================
# 기본 설정
# =========================================================

DEFAULT_EXCEL_PATH = Path(
    "26년_열린문디자인_단가표_8차(260528).xlsx"
)

DEFAULT_DB_PATH = Path(
    "price_table.db"
)

DEFAULT_REPORT_PATH = Path(
    "price_db_validation_report.json"
)

PRICE_HEADER_WORDS = {
    "단가",
    "판매가",
    "판매단가",
    "공급가",
    "공급금액",
    "금액",
    "디자인비+인쇄비",
    "인쇄비2배",
    "페이지당단가",
}

NON_SALE_HEADER_WORDS = {
    "원가",
    "마진",
    "하청가",
    "이전단가",
    "차이",
}

QUANTITY_UNITS = {
    "매",
    "부",
    "장",
    "개",
    "곽",
    "권",
    "세트",
    "조",
    "롤",
    "박스",
    "식",
}

SIZE_PATTERN = re.compile(
    r"(?<!\d)"
    r"(\d+(?:\.\d+)?)"
    r"\s*(mm|㎜|cm|㎝|m|인치|inch|in)?"
    r"\s*[*xX×ｘ]\s*"
    r"(\d+(?:\.\d+)?)"
    r"\s*(mm|㎜|cm|㎝|m|인치|inch|in)?",
    re.IGNORECASE,
)

PAPER_WEIGHT_PATTERN = re.compile(
    r"(?<!\d)(\d{2,4})\s*g\b",
    re.IGNORECASE,
)


# =========================================================
# 결과 구조
# =========================================================

@dataclass
class ValidationIssue:
    severity: str
    check_name: str
    message: str
    sheet_name: str | None = None
    row_number: int | None = None
    column_number: int | None = None
    price_item_id: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "check_name": self.check_name,
            "message": self.message,
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "column_number": self.column_number,
            "price_item_id": self.price_item_id,
        }


# =========================================================
# 텍스트 및 숫자 유틸리티
# =========================================================

def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value)
    ).strip()


def compact_text(value: Any) -> str:
    return re.sub(
        r"[\s:：·ㆍ\-_()/\[\]{},.+]",
        "",
        normalize_text(value)
    ).lower()


def to_int_price(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if value <= 0:
            return None

        return int(round(value))

    text = normalize_text(value)

    if not text or text.startswith("="):
        return None

    match = re.fullmatch(
        r"\s*₩?\s*([\d,]+(?:\.\d+)?)\s*(?:원)?\s*",
        text
    )

    if not match:
        return None

    try:
        value_number = float(
            match.group(1).replace(",", "")
        )
    except ValueError:
        return None

    if value_number <= 0:
        return None

    return int(round(value_number))


def dimension_to_mm(
    number: float,
    unit: str | None
) -> float:
    resolved = (
        unit or "mm"
    ).lower()

    if resolved in {"cm", "㎝"}:
        return number * 10

    if resolved == "m":
        return number * 1000

    if resolved in {
        "인치",
        "inch",
        "in",
    }:
        return number * 25.4

    return number


def extract_dimensions(
    text: str | None
) -> tuple[float | None, float | None]:
    if not text:
        return None, None

    match = SIZE_PATTERN.search(text)

    if not match:
        return None, None

    first = float(match.group(1))
    first_unit = match.group(2)

    second = float(match.group(3))
    second_unit = match.group(4)

    resolved_unit = (
        second_unit
        or first_unit
        or "mm"
    )

    return (
        dimension_to_mm(
            first,
            resolved_unit
        ),
        dimension_to_mm(
            second,
            resolved_unit
        ),
    )


def get_excel_cell_value(
    worksheet: Worksheet,
    row: int,
    column: int
) -> Any:
    coordinate = worksheet.cell(
        row,
        column
    ).coordinate

    for merged_range in (
        worksheet.merged_cells.ranges
    ):
        if coordinate in merged_range:
            return worksheet.cell(
                merged_range.min_row,
                merged_range.min_col
            ).value

    return worksheet.cell(
        row,
        column
    ).value


def find_price_header_above(
    worksheet: Worksheet,
    row: int,
    column: int,
    max_distance: int = 5
) -> str | None:
    for target_row in range(
        row - 1,
        max(
            0,
            row - max_distance - 1
        ),
        -1
    ):
        text = normalize_text(
            get_excel_cell_value(
                worksheet,
                target_row,
                column
            )
        )

        if not text:
            continue

        compact = compact_text(text)

        all_headers = (
            PRICE_HEADER_WORDS
            | NON_SALE_HEADER_WORDS
        )

        for header in all_headers:
            if compact == compact_text(
                header
            ):
                return header

    return None


# =========================================================
# DB
# =========================================================

def connect_database(
    database_path: Path
) -> sqlite3.Connection:
    connection = sqlite3.connect(
        database_path
    )

    connection.row_factory = (
        sqlite3.Row
    )

    return connection


def require_columns(
    connection: sqlite3.Connection
) -> None:
    existing_columns = {
        row["name"]
        for row in connection.execute(
            """
            PRAGMA table_info(
                price_items
            )
            """
        ).fetchall()
    }

    required_columns = {
        "id",
        "product_name",
        "normalized_name",
        "specification",
        "width_mm",
        "height_mm",
        "quantity",
        "unit",
        "unit_price",
        "total_price",
        "sheet_name",
        "row_number",
        "column_number",
        "review_required",
        "original_text",
    }

    missing = (
        required_columns
        - existing_columns
    )

    if missing:
        raise RuntimeError(
            "price_items 필수 컬럼이 없습니다: "
            + ", ".join(
                sorted(missing)
            )
        )


# =========================================================
# 검증
# =========================================================

def validate_database(
    excel_path: Path,
    database_path: Path
) -> tuple[
    list[ValidationIssue],
    dict[str, Any]
]:
    issues: list[ValidationIssue] = []

    workbook = load_workbook(
        excel_path,
        data_only=True,
        read_only=False
    )

    connection = connect_database(
        database_path
    )

    try:
        require_columns(
            connection
        )

        rows = connection.execute(
            """
            SELECT *
            FROM price_items
            ORDER BY id
            """
        ).fetchall()

        # -------------------------------------------------
        # 1. 기본 무결성
        # -------------------------------------------------

        for row in rows:
            price_item_id = row["id"]

            if (
                row["unit_price"] is None
                and row["total_price"] is None
            ):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="price_presence",
                        message=(
                            "unit_price와 total_price가 "
                            "모두 NULL입니다."
                        ),
                        sheet_name=row[
                            "sheet_name"
                        ],
                        row_number=row[
                            "row_number"
                        ],
                        column_number=row[
                            "column_number"
                        ],
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

            if not normalize_text(
                row["product_name"]
            ):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="product_name",
                        message=(
                            "product_name이 비어 있습니다."
                        ),
                        sheet_name=row[
                            "sheet_name"
                        ],
                        row_number=row[
                            "row_number"
                        ],
                        column_number=row[
                            "column_number"
                        ],
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

            if not normalize_text(
                row["normalized_name"]
            ):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="normalized_name",
                        message=(
                            "normalized_name이 "
                            "비어 있습니다."
                        ),
                        sheet_name=row[
                            "sheet_name"
                        ],
                        row_number=row[
                            "row_number"
                        ],
                        column_number=row[
                            "column_number"
                        ],
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

            # 수량과 단위
            if (
                row["quantity"] is not None
                and row["quantity"] <= 0
            ):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="quantity",
                        message=(
                            "수량이 0 이하입니다: "
                            f"{row['quantity']}"
                        ),
                        sheet_name=row[
                            "sheet_name"
                        ],
                        row_number=row[
                            "row_number"
                        ],
                        column_number=row[
                            "column_number"
                        ],
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

            if (
                row["unit"] is not None
                and row["unit"]
                not in QUANTITY_UNITS
                and row["unit"]
                not in {
                    "페이지",
                    "세트",
                }
            ):
                issues.append(
                    ValidationIssue(
                        severity="WARNING",
                        check_name="unit",
                        message=(
                            "알 수 없는 단위입니다: "
                            f"{row['unit']}"
                        ),
                        sheet_name=row[
                            "sheet_name"
                        ],
                        row_number=row[
                            "row_number"
                        ],
                        column_number=row[
                            "column_number"
                        ],
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

            # unit_price × quantity와 total_price 대조
            if (
                row["unit_price"] is not None
                and row["quantity"] is not None
                and row["total_price"] is not None
            ):
                expected_total = int(
                    round(
                        row["unit_price"]
                        * row["quantity"]
                    )
                )

                difference = abs(
                    expected_total
                    - row["total_price"]
                )

                tolerance = max(
                    1,
                    int(
                        row["total_price"]
                        * 0.01
                    )
                )

                if difference > tolerance:
                    issues.append(
                        ValidationIssue(
                            severity="WARNING",
                            check_name=(
                                "price_arithmetic"
                            ),
                            message=(
                                "단가×수량과 총액이 "
                                "일치하지 않습니다. "
                                f"계산={expected_total:,}, "
                                f"DB={row['total_price']:,}"
                            ),
                            sheet_name=row[
                                "sheet_name"
                            ],
                            row_number=row[
                                "row_number"
                            ],
                            column_number=row[
                                "column_number"
                            ],
                            price_item_id=(
                                price_item_id
                            ),
                        )
                    )

            # -------------------------------------------------
            # 2. 원본 Excel 셀 가격 대조
            # -------------------------------------------------

            sheet_name = row[
                "sheet_name"
            ]

            if (
                sheet_name
                not in workbook.sheetnames
            ):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="source_sheet",
                        message=(
                            "원본 Excel에 시트가 없습니다."
                        ),
                        sheet_name=sheet_name,
                        row_number=row[
                            "row_number"
                        ],
                        column_number=row[
                            "column_number"
                        ],
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )
                continue

            source_row = row[
                "row_number"
            ]

            source_column = row[
                "column_number"
            ]

            if (
                source_row is None
                or source_column is None
            ):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="source_coordinate",
                        message=(
                            "원본 행 또는 열 정보가 없습니다."
                        ),
                        sheet_name=sheet_name,
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )
                continue

            worksheet = workbook[
                sheet_name
            ]

            source_value = get_excel_cell_value(
                worksheet,
                source_row,
                source_column
            )

            source_price = to_int_price(
                source_value
            )

            db_candidate_prices = {
                value
                for value in (
                    row["unit_price"],
                    row["total_price"],
                )
                if value is not None
            }

            if source_price is not None:
                if (
                    source_price
                    not in db_candidate_prices
                ):
                    issues.append(
                        ValidationIssue(
                            severity="ERROR",
                            check_name=(
                                "source_price_match"
                            ),
                            message=(
                                "원본 셀 가격과 DB 가격이 "
                                "일치하지 않습니다. "
                                f"원본={source_price:,}, "
                                "DB="
                                + ", ".join(
                                    f"{value:,}"
                                    for value
                                    in sorted(
                                        db_candidate_prices
                                    )
                                )
                            ),
                            sheet_name=sheet_name,
                            row_number=source_row,
                            column_number=(
                                source_column
                            ),
                            price_item_id=(
                                price_item_id
                            ),
                        )
                    )

            # 원가/마진 열 여부 확인
            header = find_price_header_above(
                worksheet,
                source_row,
                source_column
            )

            if header in NON_SALE_HEADER_WORDS:
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        check_name="non_sale_price",
                        message=(
                            "판매가가 아닌 열에서 "
                            f"추출됐습니다: {header}"
                        ),
                        sheet_name=sheet_name,
                        row_number=source_row,
                        column_number=(
                            source_column
                        ),
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

            # -------------------------------------------------
            # 3. 규격 대조
            # -------------------------------------------------

            specification = normalize_text(
                row["specification"]
            )

            source_text = normalize_text(
                row["original_text"]
            )

            parsed_width, parsed_height = (
                extract_dimensions(
                    specification
                    or source_text
                )
            )

            db_width = row[
                "width_mm"
            ]

            db_height = row[
                "height_mm"
            ]

            if (
                parsed_width is not None
                and parsed_height is not None
                and db_width is not None
                and db_height is not None
            ):
                direct_match = (
                    abs(
                        parsed_width
                        - db_width
                    ) <= 1
                    and abs(
                        parsed_height
                        - db_height
                    ) <= 1
                )

                reverse_match = (
                    abs(
                        parsed_width
                        - db_height
                    ) <= 1
                    and abs(
                        parsed_height
                        - db_width
                    ) <= 1
                )

                if not (
                    direct_match
                    or reverse_match
                ):
                    issues.append(
                        ValidationIssue(
                            severity="WARNING",
                            check_name=(
                                "dimension_match"
                            ),
                            message=(
                                "규격 문자열과 DB 가로·세로가 "
                                "일치하지 않습니다. "
                                f"문자열={parsed_width:g}×"
                                f"{parsed_height:g}, "
                                f"DB={db_width:g}×"
                                f"{db_height:g}"
                            ),
                            sheet_name=sheet_name,
                            row_number=source_row,
                            column_number=(
                                source_column
                            ),
                            price_item_id=(
                                price_item_id
                            ),
                        )
                    )

            # -------------------------------------------------
            # 4. 평량을 수량으로 오인했는지 검사
            # -------------------------------------------------

            quantity = row[
                "quantity"
            ]

            if quantity is not None:
                weight_matches = {
                    float(match.group(1))
                    for match in (
                        PAPER_WEIGHT_PATTERN
                        .finditer(
                            f"{specification} "
                            f"{source_text}"
                        )
                    )
                }

                if (
                    quantity in weight_matches
                    and not row["unit"]
                ):
                    issues.append(
                        ValidationIssue(
                            severity="ERROR",
                            check_name=(
                                "paper_weight_as_quantity"
                            ),
                            message=(
                                "종이 평량을 수량으로 "
                                f"오인했을 가능성이 있습니다: "
                                f"{quantity:g}g"
                            ),
                            sheet_name=sheet_name,
                            row_number=source_row,
                            column_number=(
                                source_column
                            ),
                            price_item_id=(
                                price_item_id
                            ),
                        )
                    )

            if row[
                "review_required"
            ] != 0:
                issues.append(
                    ValidationIssue(
                        severity="WARNING",
                        check_name="review_required",
                        message=(
                            "review_required가 0이 아닙니다."
                        ),
                        sheet_name=sheet_name,
                        row_number=source_row,
                        column_number=(
                            source_column
                        ),
                        price_item_id=(
                            price_item_id
                        ),
                    )
                )

        # -------------------------------------------------
        # 5. 중복 검사
        # -------------------------------------------------

        signatures = defaultdict(
            list
        )

        for row in rows:
            signature = (
                row["normalized_name"],
                normalize_text(
                    row["specification"]
                ),
                row["quantity"],
                row["unit"],
                row["unit_price"],
                row["total_price"],
                row["sheet_name"],
                row["row_number"],
                row["column_number"],
            )

            signatures[
                signature
            ].append(
                row["id"]
            )

        for signature, ids in (
            signatures.items()
        ):
            if len(ids) <= 1:
                continue

            issues.append(
                ValidationIssue(
                    severity="WARNING",
                    check_name="duplicate_item",
                    message=(
                        "동일한 가격 항목이 중복 저장됐습니다. "
                        f"IDs={ids}"
                    ),
                    sheet_name=signature[6],
                    row_number=signature[7],
                    column_number=signature[8],
                    price_item_id=ids[0],
                )
            )

        # -------------------------------------------------
        # 요약
        # -------------------------------------------------

        severity_counts = Counter(
            issue.severity
            for issue in issues
        )

        item_counts = {
            row["normalized_name"]: row["count"]
            for row in connection.execute(
                """
                SELECT
                    normalized_name,
                    COUNT(*) AS count
                FROM price_items
                GROUP BY normalized_name
                ORDER BY count DESC
                """
            ).fetchall()
        }

        sheet_counts = {
            row["sheet_name"]: row["count"]
            for row in connection.execute(
                """
                SELECT
                    sheet_name,
                    COUNT(*) AS count
                FROM price_items
                GROUP BY sheet_name
                ORDER BY sheet_name
                """
            ).fetchall()
        }

        summary = {
            "price_items": len(rows),
            "errors": severity_counts[
                "ERROR"
            ],
            "warnings": severity_counts[
                "WARNING"
            ],
            "passed": (
                severity_counts["ERROR"]
                == 0
            ),
            "item_counts": item_counts,
            "sheet_counts": sheet_counts,
        }

        return issues, summary

    finally:
        workbook.close()
        connection.close()


# =========================================================
# 출력
# =========================================================

def print_report(
    issues: list[ValidationIssue],
    summary: dict[str, Any]
) -> None:
    print()
    print("=" * 90)
    print("PRICE TABLE DB VALIDATION")
    print("=" * 90)

    print(
        "가격 항목:",
        summary["price_items"]
    )

    print(
        "오류:",
        summary["errors"]
    )

    print(
        "경고:",
        summary["warnings"]
    )

    print(
        "최종 결과:",
        (
            "PASS"
            if summary["passed"]
            else "FAIL"
        )
    )

    print()
    print("[시트별 건수]")

    for sheet_name, count in (
        summary["sheet_counts"].items()
    ):
        print(
            f"  {sheet_name}: {count}건"
        )

    if issues:
        print()
        print("[오류 및 경고]")

        for index, issue in enumerate(
            issues,
            start=1
        ):
            location_parts: list[str] = []

            if issue.sheet_name:
                location_parts.append(
                    issue.sheet_name
                )

            if issue.row_number is not None:
                location_parts.append(
                    f"행 {issue.row_number}"
                )

            if (
                issue.column_number
                is not None
            ):
                location_parts.append(
                    f"열 {issue.column_number}"
                )

            if (
                issue.price_item_id
                is not None
            ):
                location_parts.append(
                    f"ID {issue.price_item_id}"
                )

            location = (
                " / ".join(
                    location_parts
                )
                or "위치 미확인"
            )

            print()
            print(
                f"{index}. "
                f"[{issue.severity}] "
                f"{issue.check_name}"
            )

            print(
                f"   위치: {location}"
            )

            print(
                f"   내용: {issue.message}"
            )

    print()
    print("=" * 90)


def save_json_report(
    report_path: Path,
    issues: list[ValidationIssue],
    summary: dict[str, Any]
) -> None:
    payload = {
        "summary": summary,
        "issues": [
            issue.to_dict()
            for issue in issues
        ],
    }

    report_path.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2
        ),
        encoding="utf-8"
    )


# =========================================================
# 실행
# =========================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "열린문디자인 단가표 Excel과 "
            "price_table.db를 비교 검증합니다."
        )
    )

    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="원본 단가표 Excel 경로"
    )

    parser.add_argument(
        "--db",
        type=Path,
        default=DEFAULT_DB_PATH,
        help="검증할 SQLite DB 경로"
    )

    parser.add_argument(
        "--report",
        type=Path,
        default=DEFAULT_REPORT_PATH,
        help="JSON 결과 보고서 경로"
    )

    parser.add_argument(
        "--fail-on-warning",
        action="store_true",
        help=(
            "경고가 하나라도 있으면 "
            "종료 코드 1을 반환"
        )
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if not args.excel.exists():
        raise FileNotFoundError(
            "원본 Excel 파일이 없습니다: "
            f"{args.excel.resolve()}"
        )

    if not args.db.exists():
        raise FileNotFoundError(
            "DB 파일이 없습니다: "
            f"{args.db.resolve()}"
        )

    issues, summary = (
        validate_database(
            excel_path=args.excel,
            database_path=args.db
        )
    )

    print_report(
        issues,
        summary
    )

    save_json_report(
        report_path=args.report,
        issues=issues,
        summary=summary
    )

    print(
        "JSON 보고서:",
        args.report.resolve()
    )

    should_fail = (
        summary["errors"] > 0
        or (
            args.fail_on_warning
            and summary["warnings"] > 0
        )
    )

    if should_fail:
        sys.exit(1)


if __name__ == "__main__":
    main()
